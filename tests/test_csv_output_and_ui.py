import io
import sys
import unittest

sys.path.insert(0, __import__('os').path.dirname(__import__('os').path.dirname(__file__)))
import process_lcms_data as processor


class CsvOutputAndUiTests(unittest.TestCase):
    def test_streamlit_reloads_processor_module_before_binding_new_helpers(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertIn('import importlib', source)
        self.assertIn('import process_lcms_data as processor', source)
        self.assertIn('processor = importlib.reload(processor)', source)
        self.assertIn('build_blank_mdl_evidence = processor.build_blank_mdl_evidence', source)

    def test_blank_mdl_workflow_is_visible_before_upload(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertIn("'blank_workflow_header'", source)
        self.assertIn("'zh': 'Blank/MDL 设置与计算'", source)
        self.assertIn("'blank_workflow_before_upload'", source)
        self.assertIn("st.subheader(t('blank_workflow_header', L))", source)
        self.assertIn("st.info(t('blank_workflow_before_upload', L))", source)
        self.assertIn('APP_VERSION', source)

    def test_main_interface_shows_detailed_ss_and_is_per_ms_examples(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertIn("'ss_is_example_header'", source)
        self.assertIn('ss_example_rows =', source)
        self.assertIn('is_example_rows =', source)
        self.assertIn("'d7-C12-BAC'", source)
        self.assertIn("'MS1基质加标浓度（ppb）': 4", source)
        self.assertIn("'MS2基质加标浓度（ppb）': 8", source)
        self.assertIn("'MS3基质加标浓度（ppb）': 12", source)
        self.assertIn("'d9-C10-ATMAC'", source)
        self.assertIn("'IS-A'", source)
        self.assertIn("'IS-B'", source)
        self.assertIn("'IS内标化合物名称': 'IS-A'", source)
        self.assertNotIn("'MS1加入浓度（ppb）':", source)
        self.assertIn("st.dataframe(pd.DataFrame(ss_example_rows", source)
        self.assertIn("st.dataframe(pd.DataFrame(is_example_rows", source)

    def test_export_csv_contains_summary_columns_and_numeric_preview(self):
        raw = (
            'Name,Ion,BLANK1,BLANK2,MS1,F1,F2\n'
            'C8-BAC,248>91,0.1,0.2,10,0.4,0.5\n'
            'd7-C12-BAC,311>98,0.1,0.2,4,4,4\n'
        ).encode('utf-8')
        output, _ = processor.process({
            'input_bytes': raw, 'input_file': '', 'output_format': 'csv',
            'output_file': 'processed.csv', 'sample_type': '尿液',
            'is_compounds': [], 'ss_compounds': ['d7-C12-BAC'],
            'ss_spike_concentrations': {'d7-C12-BAC': 4},
            'spike_conc_ppb': 10, 'conversion_factor': 1,
        }, return_bytes=True)
        self.assertIn('最终浓度'.encode('utf-8'), output)
        self.assertIn(b'Median (Q1-Q3)', output)
        self.assertNotIn(b'=', output)

    def test_csv_uses_selected_language_and_records_each_is_ms_addition(self):
        raw = (
            'Name,Ion,BLANK1,BLANK2,MS1,MS2,F1\n'
            'C8-BAC,248>91,0.1,0.2,10,20,0.5\n'
            'My Internal Standard,300>100,0.1,0.2,4,8,4\n'
        ).encode('utf-8')
        output, _ = processor.process({
            'input_bytes': raw, 'input_file': '', 'output_format': 'csv',
            'output_file': 'processed.csv', 'language': 'zh',
            'is_compounds': ['My Internal Standard'],
            'matrix_spike_concentrations': {
                'My Internal Standard': {'MS1': 4, 'MS2': 8},
            },
        }, return_bytes=True)

        text = output.decode('utf-8-sig')
        self.assertIn('名称,链长,DF (%),Median (Q1-Q3),MDL,MQL', text)
        self.assertIn('IS 加入浓度记录（仅记录）,MS1,MS2,是否经过 IS 校正', text)
        self.assertIn('My Internal Standard,4.0,8.0,否', text)

    def test_language_labels_are_available_in_both_languages(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertIn("'zh': '上传原始数据（XLSX 或 CSV）'", source)
        self.assertIn("'en': 'Upload Raw Data (XLSX or CSV)'", source)
        self.assertIn("'en': 'Compound Roles'", source)
        self.assertIn("'en': 'Per-compound Blank/MDL settings'", source)

    def test_english_ui_localizes_matrix_spike_table_headers_and_roles(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertIn("'ms_table_compound'", source)
        self.assertIn("'ms_table_role'", source)
        self.assertIn("role_label(role, L)", source)

    def test_raw_preview_uses_display_safe_text_to_avoid_mixed_arrow_columns(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertIn("df_raw.head(8).fillna('').astype(str)", source)

    def test_ui_keeps_one_is_correction_question(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertNotIn("'is_mode'", source)
        self.assertNotIn('is_correction_mode', source)
        self.assertNotIn('response_factors', source)
        self.assertEqual(source.count("'is_correction':"), 1)

    def test_ui_exposes_custom_ss_input_with_example_and_validation(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertIn("'custom_ss'", source)
        self.assertIn('d7-C12-BAC，4，8，12', source)
        self.assertIn('parse_ss_matrix_spike_entries(', source)
        self.assertIn('missing_custom_ss', source)
        self.assertIn('SS基质加标浓度必须填写', source)
        self.assertNotIn('这里只输入SS化合物名称，不输入浓度', source)

    def test_ui_exposes_custom_is_name_input_and_keeps_correction_separate(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertIn("'custom_is'", source)
        self.assertIn('parse_compound_name_entries(custom_is_text)', source)
        self.assertNotIn('float(custom_is.get(name, 4.0))', source)
        self.assertIn("'is_corrected': is_corrected", source)

    def test_ui_documents_all_name_separators_and_dynamic_ms_columns(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        for marker in ('英文逗号“,”', '中文逗号“，”', '英文分号“;”', '中文分号“；”', 'Tab', '换行'):
            self.assertIn(marker, source)
        self.assertIn('MS1基质加标浓度（ppb）', source)
        self.assertIn('MS2基质加标浓度（ppb）', source)
        self.assertIn('MS3基质加标浓度（ppb）', source)
        self.assertIn('实际MS列数量', source)
        self.assertIn("for index, (_, _, header) in enumerate(mss, 1)", source)

    def test_is_rows_are_not_added_to_matrix_spike_concentration_editor(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        table_region = source[source.index('ms_rows = []'):source.index('ms_table = st.data_editor')]
        self.assertNotIn("('IS', roles_for_ms['is_compounds'])", table_region)

    def test_ss_matrix_spike_cells_have_no_silent_four_ppb_default_and_are_required(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertIn("key='custom_ss_text'", source)
        self.assertIn('missing_matrix_spike_entries(', source)
        self.assertIn('not ss_spike_ready', source)
        self.assertNotIn('ss_concentrations = {name: 4.0', source)

    def test_ss_matrix_spike_values_are_entered_in_one_sidebar_text_box(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertIn('每行填写一个替代物：名称，MS1加标浓度，MS2加标浓度……', source)
        self.assertIn('d7-C12-BAC，4，8，12', source)
        self.assertIn('d9-C10-ATMAC；2；2；4', source)
        self.assertNotIn("key=f'ss_matrix_spike_{ss_index}_{ms_index}'", source)
        table_region = source[source.index('ms_rows = []'):source.index('ms_table = st.data_editor')]
        self.assertNotIn("roles_for_ms['ss_compounds']", table_region)

    def test_sidebar_shows_concise_ss_matrix_spike_instructions_without_duplicate_info_box(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertIn('每行填写一个替代物：名称，MS1加标浓度，MS2加标浓度……', source)
        self.assertIn('名称须与原始表完全一致', source)
        self.assertIn('不是原始表中的MS实测浓度', source)
        self.assertIn('实际有几个MS，就填写几个浓度', source)
        self.assertIn('支持中英文逗号、分号或Tab分隔', source)
        self.assertIn('回收率 = 原始表SS实测浓度 ÷ 填写的SS基质加标浓度 × 100%', source)
        self.assertNotIn("st.info(t('ss_input_example_body', L))", source)
        self.assertNotIn("st.markdown(f\"**{t('ss_input_example_title', L)}**\")", source)

    def test_ui_does_not_show_fixed_d7_and_d9_surrogate_concentration_boxes(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertNotIn("ss_spike_d7 = st.number_input", source)
        self.assertNotIn("ss_spike_d9 = st.number_input", source)
        self.assertNotIn("'ss_spike_d7_ppb':", source)
        self.assertNotIn("'ss_spike_d9_ppb':", source)

    def test_custom_ss_input_is_visible_in_sidebar_before_file_upload(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        sidebar_start = source.index('with st.sidebar:')
        main_region = source.index('# 主区域')
        sidebar_source = source[sidebar_start:main_region]
        self.assertIn('custom_ss_text = st.text_area(', sidebar_source)
        self.assertIn("key='custom_ss_text'", sidebar_source)

    def test_ui_preview_accepts_legacy_xls_uploads(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertIn("file_bytes.startswith(b'\\xd0\\xcf\\x11\\xe0')", source)
        self.assertIn("pd.read_excel(io.BytesIO(file_bytes), header=None, engine='xlrd')", source)

    def test_ui_shows_a_table_of_type_chain_and_role_for_detected_compounds(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertIn('compound_classification_rows', source)
        self.assertIn('classification_rows', source)
        self.assertIn('pd.DataFrame(classification_rows)', source)

    def test_ui_uses_one_compound_by_ms_table_as_the_concentration_source_of_truth(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertIn("key='compound_matrix_spike_concentration_table'", source)
        self.assertIn('matrix_spike_concentrations = {', source)
        self.assertNotIn("key=f'ss_conc_{name}'", source)
        self.assertNotIn("key=f'is_conc_{name}'", source)

    def test_ui_shows_per_compound_blank_mdl_calculation_evidence(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertIn("'calculation_evidence'", source)
        self.assertIn("'zh': '计算依据'", source)
        self.assertIn("'en': 'Calculation evidence'", source)
        self.assertIn('build_blank_mdl_evidence', source)
        self.assertIn('blank_evidence_rows', source)
        self.assertIn('degrees_of_freedom', source)
        self.assertIn("'evidence_reason_missing'", source)

    def test_ui_removes_low_spike_replicate_input(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertNotIn('low_spike_text', source)
        self.assertNotIn('parse_low_spike_entries', source)
        self.assertNotIn("'low_spike_header'", source)

    def test_classification_table_updates_after_user_confirms_is_and_ss_roles(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'streamlit_app.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertIn('compound_classification_rows(all_c, selected_is, selected_ss, compound_metadata)', source)

    def test_is_addition_concentrations_are_recorded_without_changing_final_concentration_rule(self):
        raw = (
            'Name,Ion,BLANK1,BLANK2,MS1,F1\n'
            'C8-BAC,248>91,0.1,0.2,10,0.5\n'
            'My Internal Standard,300>100,0.1,0.2,4,4\n'
        ).encode('utf-8')
        output, _ = processor.process({
            'input_bytes': raw, 'input_file': '', 'output_file': 'is_record.xlsx',
            'is_compounds': ['My Internal Standard'],
            'is_spike_concentrations': {'My Internal Standard': 4},
            'is_corrected': True,
        }, return_bytes=True)
        import openpyxl
        workbook = openpyxl.load_workbook(io.BytesIO(output), data_only=False)
        info = workbook['计算说明']
        rows = [[info.cell(r, c).value for c in range(1, 5)] for r in range(1, info.max_row + 1)]
        marker = ['IS additions, ppb  内标加入浓度（仅记录）', 'MS1', '是否IS校正', None]
        self.assertIn(marker, rows)
        self.assertEqual(rows[rows.index(marker) + 1], ['My Internal Standard', None, 'yes', None])


if __name__ == '__main__':
    unittest.main()
