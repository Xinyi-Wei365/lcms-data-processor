import io
import sys
import unittest

import openpyxl

sys.path.insert(0, __import__('os').path.dirname(__import__('os').path.dirname(__file__)))
import process_lcms_data as processor


class ConversionFactorTests(unittest.TestCase):
    def test_final_half_mdl_fallback_does_not_apply_conversion_factor_twice(self):
        raw = (
            'Name,Ion,BLANK1,BLANK2,MS1,F1\n'
            'C8-BAC,248>91,1,1,10,0.5\n'
        ).encode('utf-8')
        output, _ = processor.process({
            'input_bytes': raw,
            'input_file': '',
            'conversion_factor': 0.25,
            'spike_conc_ppb': 10,
            'is_compounds': [],
            'ss_compounds': [],
            'output_file': 'converted.xlsx',
        }, return_bytes=True)
        wb = openpyxl.load_workbook(io.BytesIO(output), data_only=False)
        final_ws = wb['Final. conc 最终计算浓度']

        # Sheet2 K is already converted to sample concentration, so the
        # non-detect fallback must reference it directly.
        self.assertIn("'Blanks_MDL 空白基质检出限'!E5", final_ws['P4'].value)
        self.assertNotIn("'Blanks_MDL 空白基质检出限'!E5*$B$38", final_ws['P4'].value)

    def test_final_df_uses_true_detection_status_not_final_numeric_values(self):
        raw = (
            'Name,Ion,BLANK1,BLANK2,MS1,F1,F2,F3\n'
            'C8-BAC,248>91,1,1,10,0.5,0.6,2\n'
        ).encode('utf-8')
        output, _ = processor.process({
            'input_bytes': raw,
            'input_file': '',
            'conversion_factor': 1,
            'spike_conc_ppb': 10,
            'is_compounds': [],
            'ss_compounds': [],
            'output_file': 'df.xlsx',
        }, return_bytes=True)
        wb = openpyxl.load_workbook(io.BytesIO(output), data_only=False)
        final_ws = wb['Final. conc 最终计算浓度']

        self.assertIn('COUNTIF(', final_ws['D4'].value)
        self.assertIn('>0', final_ws['D4'].value)
        self.assertNotIn('COUNT(P4:R4)', final_ws['D4'].value)
        info_ws = wb['计算说明']
        self.assertTrue(any('原始瓶内浓度≥瓶内MDL' in str(info_ws.cell(row, 4).value)
                            for row in range(1, info_ws.max_row + 1)))


if __name__ == '__main__':
    unittest.main()
