import io
import sys
import unittest

import openpyxl

sys.path.insert(0, __import__('os').path.dirname(__import__('os').path.dirname(__file__)))
import process_lcms_data as processor


class FinalWorkflowTests(unittest.TestCase):
    def test_nonzero_blank_mdl_uses_dynamic_one_sided_99_percent_t(self):
        mdl = processor.calculate_nonzero_blank_mdl(
            spike_values=[0.9, 1.0, 1.1, 1.0, 1.0, 0.9, 1.1],
            blank_values=[0.20, 0.21, 0.19, 0.20, 0.20, 0.21, 0.19],
        )
        self.assertAlmostEqual(mdl, 0.2256625, places=6)

    def test_df_uses_bottle_mdl_not_blank_average(self):
        raw_data = {'C8-BAC': {'B': 0.19, 'C': 0.20, 'D': 0.21, 'E': 0.25, 'F': 0.60, 'G': None}}
        blanks = [(2, 'B', 'BLANK1'), (3, 'C', 'BLANK2'), (4, 'D', 'BLANK3')]
        samples = [(5, 'E', 'F1'), (6, 'F', 'F2'), (7, 'G', 'F3')]
        cfg = {'target_compounds': ['C8-BAC'], 'conversion_factor': 1}
        rows = processor.compute_preview_summary(raw_data, blanks, samples, cfg)
        self.assertAlmostEqual(rows[0]['DF (%)'], 50.0)

    def test_metadata_override_controls_type_chain_role_and_order(self):
        metadata = {
            'Unknown analyte': {'type': 'PFAS', 'chain_length': 'C8', 'role': 'Target'},
            'Custom monitor': {'type': 'QA', 'chain_length': 'NA', 'role': 'IS'},
        }
        roles = processor.configured_compound_lists(
            {'compound_metadata': metadata}, ['Unknown analyte', 'Custom monitor'], [], []
        )
        self.assertEqual(roles[0], ['Unknown analyte'])
        self.assertEqual(roles[1], ['Custom monitor'])
        self.assertEqual(processor.compound_metadata_for('Unknown analyte', metadata)['type'], 'PFAS')

    def test_chinese_role_labels_are_normalized_before_processing(self):
        metadata = {
            'My internal': {'role': '内标'},
            'My surrogate': {'role': '替代物'},
            'My target': {'role': '目标物'},
        }
        targets, internal_standards, surrogates, _ = processor.configured_compound_lists(
            {'compound_metadata': metadata},
            ['My internal', 'My surrogate', 'My target'], [], [],
        )
        self.assertEqual(targets, ['My target'])
        self.assertEqual(internal_standards, ['My internal'])
        self.assertEqual(surrogates, ['My surrogate'])

    def test_is_is_not_in_recovery_rows_and_is_additions_are_cells_per_ms(self):
        raw = (
            'Name,Ion,BLANK1,BLANK2,MS1,MS2,F1\n'
            'C8-BAC,248>91,0.1,0.2,10,20,0.5\n'
            'My Surrogate,300>100,0.1,0.2,4,2,2\n'
            'My Internal Standard,301>101,0.1,0.2,4,8,4\n'
        ).encode('utf-8')
        out, _ = processor.process({
            'input_bytes': raw, 'input_file': '', 'output_file': 'is.xlsx',
            'is_compounds': ['My Internal Standard'], 'ss_compounds': ['My Surrogate'],
            'matrix_spike_concentrations': {
                'C8-BAC': {'MS1': 10, 'MS2': 20},
                'My Surrogate': {'MS1': 4, 'MS2': 2},
                'My Internal Standard': {'MS1': 4, 'MS2': 8},
            },
            'mdl_spike_values': {'C8-BAC': [0.9, 1, 1.1, 1, 1, 0.9, 1.1]},
        }, return_bytes=True)
        wb = openpyxl.load_workbook(io.BytesIO(out), data_only=False)
        matrix = wb[next(name for name in wb.sheetnames if name.startswith('Matrix spike'))]
        is_row = next(row for row in range(1, matrix.max_row + 1) if matrix.cell(row, 1).value == 'My Internal Standard')
        self.assertEqual(matrix.cell(is_row, 1).value, 'My Internal Standard')
        self.assertEqual(matrix.cell(is_row, 6).value, None)
        self.assertTrue(any(str(matrix.cell(row, 1).value or '').startswith('IS measured')
                            for row in range(1, matrix.max_row + 1)))
        self.assertNotIn('统计计算结果', wb.sheetnames)

    def test_english_workbook_uses_english_sheet_titles_and_final_results_summary(self):
        raw = (
            'Name,Ion,BLANK1,BLANK2,MS1,F1\n'
            'C8-DDAC,248>91,0.1,0.2,10,0.5\n'
        ).encode('utf-8')
        out, _ = processor.process({
            'input_bytes': raw, 'input_file': '', 'output_file': 'english.xlsx',
            'language': 'en', 'mdl_spike_values': {'C8-DADMAC': [0.9, 1, 1.1, 1, 1, 0.9, 1.1]},
        }, return_bytes=True)
        wb = openpyxl.load_workbook(io.BytesIO(out), data_only=False)
        self.assertIn('Descriptive statistics', wb.sheetnames)
        self.assertNotIn('描述性统计', wb.sheetnames)
        summary = wb['Descriptive statistics']
        self.assertEqual([summary.cell(3, c).value for c in range(1, 7)],
                         ['Name', 'Chain length', 'DF (%)', 'Median (Q1-Q3)', 'MDL', 'MQL'])
        self.assertEqual(summary['A4'].value, 'C8-DADMAC')


if __name__ == '__main__':
    unittest.main()
