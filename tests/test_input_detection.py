import io
import sys
import unittest

from openpyxl import Workbook

sys.path.insert(0, __import__('os').path.dirname(__import__('os').path.dirname(__file__)))
import process_lcms_data as processor


class InputDetectionTests(unittest.TestCase):
    def test_csv_bytes_are_read_and_other_analytes_are_preserved(self):
        csv_bytes = (
            'Name,Ion,F91-BLANK1,F89-MS1,F1,F2\n'
            'C8-BAC,248.2>91.0,0.1,10.0,0.2,0.3\n'
            'C12-Other,300.0>100.0,0.2,9.0,0.4,0.5\n'
        ).encode('utf-8')

        raw, blanks, mss, samples, targets, is_comps, ss_comps, all_comps = processor.read_raw(csv_bytes)

        self.assertIn('C12-Other', targets)
        self.assertEqual(len(blanks), 1)
        self.assertEqual(len(mss), 1)
        self.assertEqual(len(samples), 2)
        self.assertIn('C12-Other', raw)

    def test_xlsx_reader_uses_first_nonempty_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.title = '说明'
        ws['A1'] = 'not data'
        data = wb.create_sheet('Results')
        data.append(['Name', 'Ion', 'F91-BLANK1', 'F1'])
        data.append(['C8-BAC', '248.2>91.0', 0.1, 0.2])
        buf = io.BytesIO()
        wb.save(buf)

        raw, blanks, mss, samples, targets, _, _, _ = processor.read_raw(buf.getvalue())

        self.assertIn('C8-BAC', targets)
        self.assertEqual(len(blanks), 1)
        self.assertEqual(len(samples), 1)
        self.assertIn('C8-BAC', raw)

    def test_chain_length_is_extracted_without_renaming_ddac(self):
        self.assertEqual(processor.extract_chain_length('C12-DDAC'), 'C12')
        self.assertEqual(processor.extract_chain_length('C10-Other'), 'C10')
        self.assertIsNone(processor.extract_chain_length('Unknown analyte'))

    def test_xlsx_generic_layout_keeps_rows_after_preview_window(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Results'
        ws.append(['Name', 'Ion', 'F91-BLANK1', 'F1'])
        for index in range(30):
            ws.append([f'C{index + 1}-Other', '300>100', 0.1, 0.2])
        buf = io.BytesIO()
        wb.save(buf)

        raw, _, _, _, targets, _, _, _ = processor.read_raw(buf.getvalue())

        self.assertEqual(len(raw), 30)
        self.assertIn('C30-Other', targets)

    def test_xlsx_ignores_non_data_sheet_even_when_named_sheet1(self):
        wb = Workbook()
        notes = wb.active
        notes.title = 'Sheet1'
        notes['A1'] = 'method notes'
        data = wb.create_sheet('Export')
        data.append(['Name', 'Ion', 'F91-BLANK1', 'F1'])
        data.append(['C8-BAC', '248>91', 0.1, 0.2])
        buf = io.BytesIO()
        wb.save(buf)

        raw, _, _, _, targets, _, _, _ = processor.read_raw(buf.getvalue())

        self.assertEqual(targets, ['C8-BAC'])
        self.assertIn('C8-BAC', raw)

    def test_masshunter_style_analyte_name_and_transition_headers_are_recognized(self):
        csv_bytes = (
            'Analyte Name,Transition,Blank 1,Matrix Spike 1,Sample 01\n'
            'C8-PFAS,499>80,0.01,1.0,0.2\n'
        ).encode('utf-8')

        raw, blanks, mss, samples, targets, _, _, _ = processor.read_raw(csv_bytes)

        self.assertIn('C8-PFAS', raw)
        self.assertEqual([item[2] for item in blanks], ['Blank 1'])
        self.assertEqual([item[2] for item in mss], ['Matrix Spike 1'])
        self.assertEqual([item[2] for item in samples], ['Sample 01'])
        self.assertEqual(targets, ['C8-PFAS'])

    def test_utf16_csv_export_is_read_without_misclassifying_headers(self):
        csv_bytes = (
            'Compound Name,Transition,BLANK1,MS1,F1\n'
            'C10-PFOS,499>80,0.01,1.0,0.2\n'
        ).encode('utf-16')

        raw, blanks, mss, samples, targets, _, _, _ = processor.read_raw(csv_bytes)

        self.assertIn('C10-PFOS', raw)
        self.assertEqual(len(blanks), 1)
        self.assertEqual(len(mss), 1)
        self.assertEqual([item[2] for item in samples], ['F1'])
        self.assertEqual(targets, ['C10-PFOS'])

    def test_legacy_xls_file_is_read_and_keeps_non_qac_analytes(self):
        frame = processor.pd.DataFrame([
            ['Compound Name', 'Transition', 'BLANK1', 'MS1', 'Sample 1'],
            ['C8-PFAS', '499>80', 0.01, 1.0, 0.2],
        ])
        output = io.BytesIO()
        try:
            frame.to_excel(output, index=False, header=False, engine='xlwt')
        except (ModuleNotFoundError, ValueError):
            self.skipTest('legacy XLS writer is not installed in this runtime')

        raw, blanks, mss, samples, targets, _, _, _ = processor.read_raw(output.getvalue())

        self.assertIn('C8-PFAS', raw)
        self.assertEqual(len(blanks), 1)
        self.assertEqual(len(mss), 1)
        self.assertEqual(targets, ['C8-PFAS'])

    def test_converted_xlsx_style_underscored_headers_are_recognized(self):
        csv_bytes = (
            'Compound_Name,Ion Transition,Method Blank 1,MatrixSpike_1,Sample_01\n'
            'C8-PFAS,499>80,0.01,1.0,0.2\n'
        ).encode('utf-8')

        _, blanks, mss, samples, targets, _, _, _ = processor.read_raw(csv_bytes)

        self.assertEqual([item[2] for item in blanks], ['Method Blank 1'])
        self.assertEqual([item[2] for item in mss], ['MatrixSpike_1'])
        self.assertEqual([item[2] for item in samples], ['Sample_01'])
        self.assertEqual(targets, ['C8-PFAS'])

    def test_converted_xlsx_underscored_headers_use_the_generic_reader(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Converted export'
        ws.append(['Compound_Name', 'Ion Transition', 'Method Blank 1', 'MatrixSpike_1', 'Sample_01'])
        ws.append(['C8-PFAS', '499>80', 0.01, 1.0, 0.2])
        output = io.BytesIO()
        wb.save(output)

        raw, blanks, mss, samples, targets, _, _, _ = processor.read_raw(output.getvalue())

        self.assertIn('C8-PFAS', raw)
        self.assertEqual(len(blanks), 1)
        self.assertEqual(len(mss), 1)
        self.assertEqual([item[2] for item in samples], ['Sample_01'])
        self.assertEqual(targets, ['C8-PFAS'])


if __name__ == '__main__':
    unittest.main()
