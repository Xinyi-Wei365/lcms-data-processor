import sys
import unittest
import io

import openpyxl

sys.path.insert(0, __import__('os').path.dirname(__import__('os').path.dirname(__file__)))
import process_lcms_data as processor


class DynamicRulesTests(unittest.TestCase):
    def test_analyte_metadata_normalizes_ddac_and_extracts_type_and_chain(self):
        metadata = processor.analyte_metadata('C12-DDAC')
        self.assertEqual(metadata['name'], 'C12-DADMAC')
        self.assertEqual(metadata['type'], 'DADMAC')
        self.assertEqual(metadata['chain_length'], 'C12')

    def test_dadmac_is_recognized_as_distinct_name_without_renaming_ddac(self):
        metadata = processor.analyte_metadata('C10-DDAC')
        self.assertEqual(metadata['name'], 'C10-DADMAC')
        self.assertEqual(metadata['type'], 'DADMAC')
        targets, _, _ = processor.classify_compounds(['C10-DDAC'])
        self.assertEqual(targets, ['C10-DADMAC'])

    def test_non_qac_family_is_inferred_from_the_analyte_name_and_sorted_by_type(self):
        pfas = processor.analyte_metadata('C8-PFAS')
        pfos = processor.analyte_metadata('C10-PFOS')

        self.assertEqual(pfas['type'], 'PFAS')
        self.assertEqual(pfos['type'], 'PFOS')
        self.assertEqual(
            processor.sort_compounds(['C12-PFAS', 'C8-PFAS', 'C10-PFOS']),
            ['C8-PFAS', 'C10-PFOS', 'C12-PFAS'],
        )

    def test_role_selection_uses_all_detected_compounds(self):
        roles = processor.resolve_roles(
            ['C8-BAC', 'Surrogate-X', 'Internal-X'],
            is_compounds=['Internal-X'],
            ss_compounds=['Surrogate-X'],
        )
        self.assertEqual(roles['target_compounds'], ['C8-BAC'])
        self.assertEqual(roles['is_compounds'], ['Internal-X'])
        self.assertEqual(roles['ss_compounds'], ['Surrogate-X'])

    def test_compound_classification_rows_expose_name_type_chain_and_role(self):
        rows = processor.compound_classification_rows(
            ['C12-PFAS', 'C8-PFAS', 'Internal-X', 'Surrogate-X'],
            is_compounds=['Internal-X'],
            ss_compounds=['Surrogate-X'],
        )

        self.assertEqual(
            rows,
            [
                {'名称': 'C8-PFAS', '类型': 'PFAS', '链长': 'C8', '角色': 'Target'},
                {'名称': 'C12-PFAS', '类型': 'PFAS', '链长': 'C12', '角色': 'Target'},
                {'名称': 'Internal-X', '类型': 'Other', '链长': 'NA', '角色': 'IS'},
                {'名称': 'Surrogate-X', '类型': 'Other', '链长': 'NA', '角色': 'SS'},
            ],
        )

    def test_role_keywords_detect_custom_is_and_ss_names(self):
        targets, is_compounds, ss_compounds = processor.classify_compounds([
            'C8-PFAS', 'PFAS internal standard', 'PFAS surrogate'
        ])
        self.assertEqual(targets, ['C8-PFAS'])
        self.assertEqual(is_compounds, ['PFAS internal standard'])
        self.assertEqual(ss_compounds, ['PFAS surrogate'])

    def test_mdl_formula_uses_signal_to_noise_for_blank_zero(self):
        formula = processor.mdl_formula(
            'C12-Other',
            'B4:G4',
            {'mdl_overrides': {'C12-Other': {
                'blank_zero': True,
                'calibration_concentration': 1.0,
                'signal_to_noise': 30.0,
            }}}
        )
        self.assertEqual(formula, '=3*1.0/30.0')

    def test_mdl_report_formula_applies_conversion_factor_once(self):
        formula = processor.mdl_report_formula(
            'C12-Other', 'J5',
            {'conversion_factor': 0.25, 'mdl_overrides': {
                'C12-Other': {
                    'blank_zero': True,
                    'calibration_concentration': 1.0,
                    'signal_to_noise': 30.0,
                }
            }}
        )
        self.assertEqual(formula, '=3*1.0/30.0*0.25')

    def test_mdl_formula_keeps_blank_standard_deviation_rule_by_default(self):
        formula = processor.mdl_formula('C8-BAC', 'B4:G4', {})
        self.assertEqual(formula, '=AVERAGE(B4:G4)+T.INV(0.99,COUNT(B4:G4)-1)*STDEV.S(B4:G4)')

    def test_selected_ss_concentration_is_resolved_by_exact_name(self):
        cfg = {'ss_spike_concentrations': {'Surrogate-X': 2.0}}
        self.assertEqual(processor.resolve_ss_spike('Surrogate-X', cfg), 2.0)

    def test_custom_ss_entries_parse_name_and_individual_spike_concentration(self):
        entries, errors = processor.parse_custom_ss_entries(
            'd7-C12-BAC, 4\nMy Surrogate, 2.5'
        )
        self.assertEqual(errors, [])
        self.assertEqual(entries, {'d7-C12-BAC': 4.0, 'My Surrogate': 2.5})

    def test_custom_ss_entries_reject_invalid_lines(self):
        entries, errors = processor.parse_custom_ss_entries('d7-C12-BAC\nX, 0')
        self.assertEqual(entries, {})
        self.assertEqual(len(errors), 2)

    def test_compound_name_input_accepts_all_documented_separators(self):
        names = processor.parse_compound_name_entries(
            'SS-A, SS-B，SS-C;SS-D；SS-E\tSS-F\nSS-G'
        )
        self.assertEqual(
            names,
            ['SS-A', 'SS-B', 'SS-C', 'SS-D', 'SS-E', 'SS-F', 'SS-G'],
        )

    def test_compound_name_input_removes_duplicates_without_reordering(self):
        names = processor.parse_compound_name_entries('IS-A，IS-B\nIS-A')
        self.assertEqual(names, ['IS-A', 'IS-B'])

    def test_missing_ss_matrix_spikes_reports_every_unfilled_ms_cell(self):
        missing = processor.missing_matrix_spike_entries(
            ['SS-A', 'SS-B'], ['MS1', 'MS2'],
            {
                'SS-A': {'MS1': 4, 'MS2': None},
                'SS-B': {'MS1': 0, 'MS2': 8},
            },
        )
        self.assertEqual(missing, [('SS-A', 'MS2'), ('SS-B', 'MS1')])

    def test_missing_ss_matrix_spikes_accepts_positive_finite_values(self):
        missing = processor.missing_matrix_spike_entries(
            ['SS-A'], ['MS1', 'MS2'], {'SS-A': {'MS1': 4, 'MS2': 8}}
        )
        self.assertEqual(missing, [])

    def test_ss_matrix_spike_text_parses_name_and_all_ms_values(self):
        entries, errors = processor.parse_ss_matrix_spike_entries(
            'd7-C12-BAC，4，8，12\n'
            'd9-C10-ATMAC;2;2;4\n'
            'SS-X\t1\t3\t5',
            ['MS1', 'MS2', 'MS3'],
        )
        self.assertEqual(errors, [])
        self.assertEqual(entries['d7-C12-BAC'], {'MS1': 4.0, 'MS2': 8.0, 'MS3': 12.0})
        self.assertEqual(entries['d9-C10-ATMAC'], {'MS1': 2.0, 'MS2': 2.0, 'MS3': 4.0})
        self.assertEqual(entries['SS-X'], {'MS1': 1.0, 'MS2': 3.0, 'MS3': 5.0})

    def test_ss_matrix_spike_text_rejects_wrong_count_and_nonpositive_values(self):
        entries, errors = processor.parse_ss_matrix_spike_entries(
            'SS-A,4,8\nSS-B；2；0；4', ['MS1', 'MS2', 'MS3']
        )
        self.assertEqual(entries, {})
        self.assertEqual(len(errors), 2)
        self.assertIn('第1行', errors[0])
        self.assertIn('需要3个浓度', errors[0])
        self.assertIn('第2行', errors[1])
        self.assertIn('必须为大于0的数字', errors[1])

    def test_ss_matrix_spike_text_ignores_trailing_separators(self):
        entries, errors = processor.parse_ss_matrix_spike_entries(
            'd9-C10-ATMAC, 4, 4;\n'
            'd7-C12-BAC，4，4，\n'
            'SS-X；2；3；',
            ['MS1', 'MS2'],
        )
        self.assertEqual(errors, [])
        self.assertEqual(entries['d9-C10-ATMAC'], {'MS1': 4.0, 'MS2': 4.0})
        self.assertEqual(entries['d7-C12-BAC'], {'MS1': 4.0, 'MS2': 4.0})
        self.assertEqual(entries['SS-X'], {'MS1': 2.0, 'MS2': 3.0})

    def test_custom_ss_is_moved_to_recovery_section_and_uses_its_own_spike(self):
        raw = (
            'Name,Ion,BLANK1,BLANK2,MS1,MS2,F1\n'
            'C8-BAC,248>91,0.1,0.2,10,10,0.5\n'
            'My Surrogate,300>100,0.1,0.2,2,4,2\n'
        ).encode('utf-8')
        output, _ = processor.process({
            'input_bytes': raw, 'input_file': '', 'output_file': 'custom_ss.xlsx',
            'is_compounds': [], 'ss_compounds': ['My Surrogate'],
            'ss_spike_concentrations': {'My Surrogate': 2},
        }, return_bytes=True)
        workbook = openpyxl.load_workbook(io.BytesIO(output), data_only=False)
        matrix_sheet = workbook[next(name for name in workbook.sheetnames if name.startswith('Matrix spike'))]
        row = next(row for row in range(1, matrix_sheet.max_row + 1)
                   if matrix_sheet.cell(row, 1).value == 'My Surrogate')
        # The SS spike concentration is a processing parameter, not an
        # additional worksheet column.  For two MS replicates, SS recovery
        # values are F/G and the requested summary columns are I/J/K.
        self.assertNotIn('SS spike conc.', [cell.value for cell in matrix_sheet[1]])
        self.assertEqual(matrix_sheet.cell(row, 6).value, 100)
        self.assertEqual(matrix_sheet.cell(row, 7).value, 200)
        self.assertEqual(matrix_sheet.cell(row, 9).value, '=ROUND(AVERAGE(F6:G6),0)')
        self.assertEqual(matrix_sheet.cell(row, 10).value, '=ROUND(STDEV(F6:G6),0)')
        self.assertEqual(matrix_sheet.cell(row, 11).value, '=ROUND(J6/SQRT(COUNT(F6:G6)),0)')

    def test_each_matrix_spike_column_uses_its_own_configured_spike_concentration(self):
        raw = (
            'Name,Ion,BLANK1,BLANK2,MS1,MS2,F1\n'
            'C8-BAC,248>91,0.1,0.2,10,20,0.5\n'
            'My Surrogate,300>100,0.1,0.2,4,4,2\n'
            'My Internal Standard,301>101,0.1,0.2,4,4,4\n'
        ).encode('utf-8')
        output, _ = processor.process({
            'input_bytes': raw, 'input_file': '', 'output_file': 'two_ms.xlsx',
            'is_compounds': ['My Internal Standard'],
            'is_spike_concentrations': {'My Internal Standard': 4},
            'ss_compounds': ['My Surrogate'],
            'ss_spike_concentrations': {'My Surrogate': 2},
            'matrix_spike_concentrations': {'MS1': 10, 'MS2': 20},
        }, return_bytes=True)
        workbook = openpyxl.load_workbook(io.BytesIO(output), data_only=False)
        matrix_sheet = workbook[next(name for name in workbook.sheetnames if name.startswith('Matrix spike'))]
        target_row = next(row for row in range(1, matrix_sheet.max_row + 1)
                          if matrix_sheet.cell(row, 1).value == 'C8-BAC')
        ss_row = next(row for row in range(1, matrix_sheet.max_row + 1)
                      if matrix_sheet.cell(row, 1).value == 'My Surrogate')
        # Target recovery uses MS1=10 ppb and MS2=20 ppb.  SS instead uses
        # its own 2 ppb concentration for both columns.
        self.assertEqual([matrix_sheet.cell(target_row, c).value for c in (6, 7)], [100, 100])
        self.assertEqual([matrix_sheet.cell(ss_row, c).value for c in (6, 7)], [200, 200])
        info = workbook['计算说明']
        rows = [[info.cell(r, c).value for c in range(1, 5)] for r in range(1, info.max_row + 1)]
        self.assertTrue(any(row[0] == 'Matrix spike recovery' for row in rows))

    def test_compound_by_ms_concentrations_apply_to_target_ss_and_is_records(self):
        raw = (
            'Name,Ion,BLANK1,BLANK2,MS1,MS2,MS3,F1\n'
            'C8-BAC,248>91,0.1,0.2,10,20,5,0.5\n'
            'My Surrogate,300>100,0.1,0.2,4,2,1,2\n'
            'My Internal Standard,301>101,0.1,0.2,4,8,2,4\n'
        ).encode('utf-8')
        concentrations = {
            'C8-BAC': {'MS1': 10, 'MS2': 20, 'MS3': 5},
            'My Surrogate': {'MS1': 4, 'MS2': 2, 'MS3': 1},
            'My Internal Standard': {'MS1': 4, 'MS2': 8, 'MS3': 2},
        }
        output, _ = processor.process({
            'input_bytes': raw, 'input_file': '', 'output_file': 'compound_ms.xlsx',
            'is_compounds': ['My Internal Standard'],
            'ss_compounds': ['My Surrogate'],
            'matrix_spike_concentrations': concentrations,
        }, return_bytes=True)
        workbook = openpyxl.load_workbook(io.BytesIO(output), data_only=False)
        matrix_sheet = workbook[next(name for name in workbook.sheetnames if name.startswith('Matrix spike'))]
        target_row = next(row for row in range(1, matrix_sheet.max_row + 1)
                          if matrix_sheet.cell(row, 1).value == 'C8-BAC')
        ss_row = next(row for row in range(1, matrix_sheet.max_row + 1)
                      if matrix_sheet.cell(row, 1).value == 'My Surrogate')
        # Target and SS each use their own row-by-MS values: all three are 100%.
        self.assertEqual([matrix_sheet.cell(target_row, c).value for c in (7, 8, 9)], [100, 100, 100])
        self.assertEqual([matrix_sheet.cell(ss_row, c).value for c in (7, 8, 9)], [100, 100, 100])
        info = workbook['计算说明']
        rows = [[info.cell(r, c).value for c in range(1, 5)] for r in range(1, info.max_row + 1)]
        header_row = next(row for row in rows if row[:4] == ['IS measured concentrations, ppb  IS实测浓度（来源于原始MS列）', 'MS1', 'MS2', 'MS3'])
        value_index = rows.index(header_row) + 1
        self.assertEqual(rows[value_index][:4], ['My Internal Standard', 4, 8, 2])

    def test_non_qac_csv_end_to_end_keeps_type_order_and_blank_zero_snr_mdl(self):
        raw = (
            'Compound_Name,Ion Transition,BLANK1,BLANK2,MatrixSpike_1,Sample_01,Sample_02\n'
            'C12-PFAS,499>80,0.1,0.2,10,0.4,0.5\n'
            'C8-PFAS,499>80,0,0,10,0.2,0.3\n'
            'PFAS Internal Standard,500>81,0.1,0.2,4,4,4\n'
            'My Surrogate,501>82,0.1,0.2,2,2,4\n'
        ).encode('utf-8')
        output, _ = processor.process({
            'input_bytes': raw, 'input_file': '', 'output_file': 'pfas.xlsx',
            'is_compounds': ['PFAS Internal Standard'],
            'ss_compounds': ['My Surrogate'],
            'ss_spike_concentrations': {'My Surrogate': 2},
            'mdl_overrides': {'C8-PFAS': {
                'blank_zero': True, 'calibration_concentration': 1, 'signal_to_noise': 20,
            }},
        }, return_bytes=True)
        workbook = openpyxl.load_workbook(io.BytesIO(output), data_only=False)
        blank_sheet = workbook['Blanks_MDL 空白基质检出限']
        summary_sheet = workbook['描述性统计']
        matrix_sheet = workbook[next(name for name in workbook.sheetnames if name.startswith('Matrix spike'))]

        self.assertEqual([summary_sheet.cell(row, 1).value for row in (4, 5)], ['C8-PFAS', 'C12-PFAS'])
        self.assertEqual([summary_sheet.cell(row, 2).value for row in (4, 5)], ['C8', 'C12'])
        c8_blank_row = next(row for row in range(1, blank_sheet.max_row + 1)
                            if blank_sheet.cell(row, 1).value == 'C8-PFAS')
        self.assertEqual(blank_sheet.cell(c8_blank_row, 6).value, '=3*1.0/20.0')
        ss_row = next(row for row in range(1, matrix_sheet.max_row + 1)
                      if matrix_sheet.cell(row, 1).value == 'My Surrogate')
        self.assertEqual(matrix_sheet.cell(ss_row, 5).value, 100)

    def test_unknown_compounds_are_sorted_by_chain_then_name(self):
        targets, _, _ = processor.classify_compounds(['C12-PFOS', 'C8-PFOS', 'C10-Other'])
        self.assertEqual(targets, ['C8-PFOS', 'C10-Other', 'C12-PFOS'])

    def test_non_isotope_cooh_metabolite_is_not_forced_to_is(self):
        targets, is_compounds, _ = processor.classify_compounds(['C10-BAC-COOH', 'C10-BAC-COOH[C13]'])
        self.assertIn('C10-BAC-COOH', targets)
        self.assertNotIn('C10-BAC-COOH', is_compounds)
        self.assertIn('C10-BAC-COOH[C13]', is_compounds)

    def test_duplicate_detected_names_are_removed(self):
        targets, is_compounds, ss_compounds = processor.classify_compounds(
            ['C8-BAC', 'C8-BAC', 'd7-C12-BAC', 'd7-C12-BAC']
        )
        self.assertEqual(targets, ['C8-BAC'])
        self.assertEqual(is_compounds, [])
        self.assertEqual(ss_compounds, ['d7-C12-BAC'])

    def test_snr_path_detection_accepts_only_zero_or_empty_blank_cells(self):
        raw_data = {
            'Zero': {'B': 0, 'C': 0.0, 'D': '0'},
            'Mixed': {'B': 0, 'C': 0.1, 'D': 0},
            'Missing': {'B': 0, 'C': None, 'D': 0},
            'All empty': {'B': None, 'C': '', 'D': None},
            'Text ND': {'B': 0, 'C': 'ND', 'D': 0},
        }
        blanks = [(2, 'B', 'blank_1'), (3, 'C', 'blank_2'), (4, 'D', 'blank_3')]
        self.assertEqual(
            processor.detect_blank_zero_compounds(raw_data, blanks),
            ['Zero', 'Missing', 'All empty'],
        )

    def test_zero_or_empty_blank_series_uses_manual_snr_mdl_and_mql(self):
        cfg = {'mdl_overrides': {'Analyte': {
            'blank_zero': True,
            'calibration_concentration': 1,
            'signal_to_noise': 10,
        }}}
        evidence = processor.build_blank_mdl_evidence('Analyte', [None, 0, ''], cfg)
        self.assertEqual(evidence['status'], 'blank_zero')
        self.assertTrue(evidence['ready'])
        self.assertAlmostEqual(evidence['mdl'], 0.3)
        self.assertEqual(processor.mql_formula('Analyte', 'B2:D2', cfg), '=10*1.0/10.0')

    def test_nonzero_blank_values_remain_automatic_when_other_cells_are_empty(self):
        evidence = processor.build_blank_mdl_evidence('Analyte', [None, 0.1, 0.2, ''], {})
        self.assertEqual(evidence['status'], 'blank_nonzero')
        self.assertTrue(evidence['ready'])

    def test_text_nd_is_not_silently_treated_as_empty_or_zero(self):
        evidence = processor.build_blank_mdl_evidence('Analyte', [0, 'ND', 0], {})
        self.assertEqual(evidence['status'], 'incomplete')
        self.assertFalse(evidence['ready'])

    def test_blank_zero_mdl_requires_manual_snr_override(self):
        with self.assertRaises(ValueError):
            processor.validate_blank_zero_mdl('Zero', [0, 0, 0], {})
        self.assertIsNone(processor.validate_blank_zero_mdl(
            'Zero', [0, 0, 0],
            {'mdl_overrides': {'Zero': {
                'blank_zero': True,
                'calibration_concentration': 1,
                'signal_to_noise': 30,
            }}}
        ))

    def test_blank_zero_configuration_validation_covers_all_target_compounds(self):
        raw_data = {'Zero': {'B': 0, 'C': 0}, 'Nonzero': {'B': 0, 'C': 0.1}}
        blanks = [(2, 'B', 'blank_1'), (3, 'C', 'blank_2')]
        with self.assertRaises(ValueError):
            processor.validate_blank_zero_configuration(raw_data, blanks, ['Zero'], {})
        processor.validate_blank_zero_configuration(
            raw_data, blanks, ['Zero'],
            {'mdl_overrides': {'Zero': {
                'blank_zero': True,
                'calibration_concentration': 1,
                'signal_to_noise': 30,
            }}}
        )


if __name__ == '__main__':
    unittest.main()
