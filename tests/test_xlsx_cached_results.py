import io
import sys
import unittest

import openpyxl

sys.path.insert(0, __import__('os').path.dirname(__import__('os').path.dirname(__file__)))
import process_lcms_data as processor


class XlsxCachedResultsTests(unittest.TestCase):
    def make_output(self):
        raw = (
            'Name,Ion,BLANK1,BLANK2,BLANK3,MS1,MS2,F1,F2,F3\n'
            'C8-BAC,248>91,1,1,1,10,20,0.5,2,3\n'
            'My Internal,300>100,0,0,0,1.8,0.9,1,1,1\n'
        ).encode('utf-8')
        return processor.process({
            'input_bytes': raw,
            'input_file': '',
            'output_file': 'cached.xlsx',
            'target_compounds': ['C8-BAC'],
            'is_compounds': ['My Internal'],
            'ss_compounds': [],
            'conversion_factor': 1,
            'matrix_spike_concentrations': {'C8-BAC': {'MS1': 10, 'MS2': 20}},
        }, return_bytes=True)[0]

    def test_is_measured_ms_concentrations_are_copied_from_raw_input(self):
        output = self.make_output()
        wb = openpyxl.load_workbook(io.BytesIO(output), data_only=False)
        ws = wb[next(name for name in wb.sheetnames if name.startswith('Matrix spike'))]
        row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == 'My Internal')
        self.assertEqual(ws.cell(row, 2).value, 1.8)
        self.assertEqual(ws.cell(row, 3).value, 0.9)
        self.assertTrue(any('IS measured' in str(ws.cell(r, 1).value or '')
                            for r in range(1, ws.max_row + 1)))

    def test_matrix_spike_average_sd_and_se_have_cached_results(self):
        output = self.make_output()
        formula_wb = openpyxl.load_workbook(io.BytesIO(output), data_only=False)
        value_wb = openpyxl.load_workbook(io.BytesIO(output), data_only=True)
        name = next(sheet for sheet in formula_wb.sheetnames if sheet.startswith('Matrix spike'))
        formula_ws = formula_wb[name]
        value_ws = value_wb[name]
        row = next(r for r in range(1, formula_ws.max_row + 1)
                   if formula_ws.cell(r, 1).value == 'C8-BAC')
        average_col = next(c for c in range(1, formula_ws.max_column + 1)
                           if formula_ws.cell(1, c).value == 'average')
        self.assertTrue(str(formula_ws.cell(row, average_col).value).startswith('='))
        self.assertEqual(value_ws.cell(row, average_col).value, 100)
        self.assertEqual(value_ws.cell(row, average_col + 1).value, 0)
        self.assertEqual(value_ws.cell(row, average_col + 2).value, 0)

    def test_formula_cells_keep_formulas_and_have_cached_numeric_results(self):
        output = self.make_output()
        formula_wb = openpyxl.load_workbook(io.BytesIO(output), data_only=False)
        value_wb = openpyxl.load_workbook(io.BytesIO(output), data_only=True)

        blank_name = next(name for name in formula_wb.sheetnames if name.startswith('Blanks_MDL'))
        final_name = next(name for name in formula_wb.sheetnames if name.startswith('Final. conc'))
        blank_formula = formula_wb[blank_name]
        blank_value = value_wb[blank_name]
        final_formula = formula_wb[final_name]
        final_value = value_wb[final_name]

        mdl_col = next(c for c in range(1, blank_formula.max_column + 1)
                       if blank_formula.cell(2, c).value == 'MDL')
        half_col = next(c for c in range(1, blank_formula.max_column + 1)
                        if blank_formula.cell(2, c).value == '1/2 MDL')
        self.assertTrue(str(blank_formula.cell(5, mdl_col).value).startswith('='))
        self.assertAlmostEqual(blank_value.cell(5, mdl_col).value, 1.0)
        self.assertAlmostEqual(blank_value.cell(5, half_col).value, 0.5)
        self.assertTrue(str(final_formula['P4'].value).startswith('='))
        self.assertAlmostEqual(final_value['P4'].value, 0.5)
        self.assertAlmostEqual(final_value['Q4'].value, 1.0)
        self.assertAlmostEqual(final_value['R4'].value, 2.0)
        self.assertAlmostEqual(final_value['D4'].value, 1.0)
        self.assertAlmostEqual(final_value['E4'].value, (0.5 + 1 + 2) / 3)

    def test_online_preview_and_cached_excel_use_the_same_python_results(self):
        raw_bytes = (
            'Name,Ion,BLANK1,BLANK2,BLANK3,MS1,MS2,F1,F2,F3\n'
            'C8-BAC,248>91,1,1,1,10,20,0.5,2,3\n'
        ).encode('utf-8')
        cfg = {
            'input_bytes': raw_bytes, 'input_file': '', 'output_file': 'same-results.xlsx',
            'target_compounds': ['C8-BAC'], 'is_compounds': [], 'ss_compounds': [],
            'conversion_factor': 1,
            'matrix_spike_concentrations': {'C8-BAC': {'MS1': 10, 'MS2': 20}},
        }
        raw_data, blanks, _, samples, *_ = processor.read_raw(raw_bytes)
        preview = processor.compute_preview_final_table(raw_data, blanks, samples, cfg)[0]
        summary = processor.compute_preview_summary(raw_data, blanks, samples, cfg)[0]
        output, _ = processor.process(cfg, return_bytes=True)
        workbook = openpyxl.load_workbook(io.BytesIO(output), data_only=True)
        final = workbook[next(name for name in workbook.sheetnames if name.startswith('Final. conc'))]
        descriptive = workbook[next(name for name in workbook.sheetnames
                                    if name.startswith('Descriptive') or name == '描述性统计')]

        self.assertEqual([preview['F1'], preview['F2'], preview['F3']],
                         [final['P4'].value, final['Q4'].value, final['R4'].value])
        self.assertEqual(summary['DF (%)'], descriptive['C4'].value)
        self.assertEqual(summary['Median (Q1-Q3)'], descriptive['D4'].value)
        self.assertEqual(summary['MDL'], descriptive['E4'].value)
        self.assertEqual(summary['MQL'], descriptive['F4'].value)

    def test_df_uses_old_template_numeric_count_over_all_sample_columns(self):
        raw = (
            'Name,Ion,BLANK1,BLANK2,MS1,F1,F2,F3\n'
            'C8-BAC,248>91,1,1,10,0.5,2,\n'
        ).encode('utf-8')
        cfg = {
            'input_bytes': raw, 'input_file': '', 'output_file': 'old-df.xlsx',
            'target_compounds': ['C8-BAC'], 'is_compounds': [], 'ss_compounds': [],
            'conversion_factor': 1,
            'matrix_spike_concentrations': {'C8-BAC': {'MS1': 10}},
        }
        output, _ = processor.process(cfg, return_bytes=True)
        formula_wb = openpyxl.load_workbook(io.BytesIO(output), data_only=False)
        value_wb = openpyxl.load_workbook(io.BytesIO(output), data_only=True)
        final_name = next(name for name in formula_wb.sheetnames if name.startswith('Final. conc'))
        self.assertEqual(formula_wb[final_name]['D4'].value, '=COUNT(P4:R4)/COLUMNS(P4:R4)')
        self.assertAlmostEqual(value_wb[final_name]['D4'].value, 2 / 3)

    def test_old_template_statistics_require_df_above_fifty_percent(self):
        raw = (
            'Name,Ion,BLANK1,BLANK2,MS1,F1,F2,F3\n'
            'C8-BAC,248>91,1,1,10,2,,\n'
        ).encode('utf-8')
        cfg = {
            'input_bytes': raw, 'input_file': '', 'output_file': 'old-stat-threshold.xlsx',
            'target_compounds': ['C8-BAC'], 'is_compounds': [], 'ss_compounds': [],
            'conversion_factor': 1,
            'matrix_spike_concentrations': {'C8-BAC': {'MS1': 10}},
        }
        output, _ = processor.process(cfg, return_bytes=True)
        formula_wb = openpyxl.load_workbook(io.BytesIO(output), data_only=False)
        value_wb = openpyxl.load_workbook(io.BytesIO(output), data_only=True)
        final_name = next(name for name in formula_wb.sheetnames if name.startswith('Final. conc'))
        final_formula = formula_wb[final_name]
        final_value = value_wb[final_name]
        for column in range(5, 14):
            self.assertIn('IF(D4>50%', final_formula.cell(4, column).value)
            self.assertEqual(final_value.cell(4, column).value, 'NC')
        descriptive_name = next(name for name in formula_wb.sheetnames
                                if name.startswith('Descriptive') or name == '描述性统计')
        self.assertEqual(value_wb[descriptive_name]['D4'].value, 'NC')

    def test_zero_or_empty_blank_path_caches_zero_average_and_snr_results(self):
        raw = (
            'Name,Ion,BLANK1,BLANK2,MS1,F1\n'
            'C10-ATMAC,200>100,,,10,2\n'
        ).encode('utf-8')
        output, _ = processor.process({
            'input_bytes': raw, 'input_file': '', 'output_file': 'zero-empty.xlsx',
            'target_compounds': ['C10-ATMAC'], 'is_compounds': [], 'ss_compounds': [],
            'conversion_factor': 1,
            'matrix_spike_concentrations': {'C10-ATMAC': {'MS1': 10}},
            'mdl_overrides': {'C10-ATMAC': {
                'blank_zero': True, 'calibration_concentration': 1, 'signal_to_noise': 10,
            }},
        }, return_bytes=True)
        formula_wb = openpyxl.load_workbook(io.BytesIO(output), data_only=False)
        value_wb = openpyxl.load_workbook(io.BytesIO(output), data_only=True)
        name = next(name for name in formula_wb.sheetnames if name.startswith('Blanks_MDL'))
        average_col = next(c for c in range(1, formula_wb[name].max_column + 1)
                           if formula_wb[name].cell(2, c).value == 'procedural blank average')
        mdl_col = next(c for c in range(1, formula_wb[name].max_column + 1)
                       if formula_wb[name].cell(2, c).value == 'MDL')
        half_col = next(c for c in range(1, formula_wb[name].max_column + 1)
                        if formula_wb[name].cell(2, c).value == '1/2 MDL')
        self.assertEqual(formula_wb[name].cell(5, average_col).value, '=0')
        self.assertEqual(value_wb[name].cell(5, average_col).value, 0)
        self.assertAlmostEqual(value_wb[name].cell(5, mdl_col).value, 0.3)
        self.assertAlmostEqual(value_wb[name].cell(5, half_col).value, 0.15)

    def test_mdl_formulas_use_legacy_excel_and_wps_compatible_functions(self):
        output = self.make_output()
        workbook = openpyxl.load_workbook(io.BytesIO(output), data_only=False)
        blank_name = next(name for name in workbook.sheetnames if name.startswith('Blanks_MDL'))
        worksheet = workbook[blank_name]
        mdl_col = next(c for c in range(1, worksheet.max_column + 1)
                       if worksheet.cell(2, c).value == 'MDL')
        formula = worksheet.cell(5, mdl_col).value
        self.assertIn('TINV(0.02,', formula)
        self.assertIn('STDEV(', formula)
        self.assertNotIn('T.INV', formula)
        self.assertNotIn('STDEV.S', formula)


if __name__ == '__main__':
    unittest.main()
