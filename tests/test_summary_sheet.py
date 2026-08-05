import io
import sys
import unittest

import openpyxl

sys.path.insert(0, __import__('os').path.dirname(__import__('os').path.dirname(__file__)))
import process_lcms_data as processor


class SummarySheetTests(unittest.TestCase):
    def test_output_contains_formula_driven_descriptive_summary(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'demo_urine_qac_masshunter.xlsx'), 'rb') as source:
            raw = source.read()
        output, _ = processor.process({
            'input_bytes': raw,
            'input_file': '',
            'sample_type': '尿液',
            'sample_volume_ml': 2,
            'final_volume_ml': 0.5,
            'extra_dilution': 1,
            'conversion_factor': 1,
            'spike_conc_ppb': 10,
            'ss_compounds': ['d7-C12-BAC', 'd9-C10-ATMAC'],
            'ss_spike_concentrations': {'d7-C12-BAC': 4, 'd9-C10-ATMAC': 4},
            'mql_factor': 3.333333333,
            'output_file': 'summary.xlsx',
        }, return_bytes=True)
        wb = openpyxl.load_workbook(io.BytesIO(output), data_only=False)
        self.assertIn('\u63cf\u8ff0\u6027\u7edf\u8ba1', wb.sheetnames)
        ws = wb['\u63cf\u8ff0\u6027\u7edf\u8ba1']
        self.assertEqual([ws.cell(3, col).value for col in range(1, 7)],
                         ['名称', '链长', 'DF (%)', 'Median (Q1-Q3)', 'MDL', 'MQL'])
        self.assertTrue(str(ws['C4'].value).startswith('='))
        self.assertTrue(str(ws['D4'].value).startswith('='))
        self.assertTrue(str(ws['E4'].value).startswith('='))
        self.assertTrue(str(ws['F4'].value).startswith('='))
        self.assertTrue(str(ws['E4'].value).startswith('='))
        self.assertIn('ROUND', str(ws['F4'].value))
        self.assertIn('ROUND', str(ws['C4'].value))
        self.assertIn('Blanks_MDL', str(ws['F4'].value))
        self.assertNotIn('>50%', str(ws['D4'].value))
        self.assertIn('不受 DF', str(ws['A2'].value))
        self.assertEqual(ws['B4'].value, 'C8')
        self.assertIn('*100', str(ws['C4'].value))
        self.assertEqual(ws['C4'].number_format, 'General')
        self.assertEqual(ws['E4'].number_format, 'General')
        self.assertEqual(ws['F4'].number_format, 'General')
        blanks = wb['Blanks_MDL 空白基质检出限']
        self.assertNotIn('>50%', str(blanks['I1'].value))

    def test_legacy_final_statistics_are_not_gated_by_df(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'demo_urine_qac_masshunter.xlsx'), 'rb') as source:
            raw = source.read()
        output, _ = processor.process({'input_bytes': raw, 'input_file': '', 'output_file': 'final.xlsx'}, return_bytes=True)
        wb = openpyxl.load_workbook(io.BytesIO(output), data_only=False)
        ws = wb[next(s for s in wb.sheetnames if s.startswith('Final. conc'))]
        self.assertNotIn('>50%', str(ws['E4'].value))

    def test_stats_source_sheet_does_not_claim_a_df_over_50_gate(self):
        with open(__import__('os').path.join(__import__('os').path.dirname(__file__), '..', 'process_lcms_data.py'), encoding='utf-8-sig') as handle:
            source = handle.read()
        self.assertNotIn('当DF>50%时ND替换', source)


if __name__ == '__main__':
    unittest.main()
