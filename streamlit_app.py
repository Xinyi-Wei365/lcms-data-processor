#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
LC-MS/MS 数据处理智能体 — Streamlit 可视化界面
"""

import streamlit as st
import pandas as pd
import tempfile
import os
import io
from process_lcms_data import (
    process, read_raw, classify_compounds, resolve_roles,
    compute_preview_summary, compute_preview_final_table, detect_blank_zero_compounds,
    parse_custom_ss_entries, compound_classification_rows,
)

try:
    from process_lcms_data import validate_input_layout
except ImportError:
    # Keep the app bootable while Streamlit Cloud refreshes an older module cache.
    def validate_input_layout(blanks, mss, samps, target_compounds, is_compounds, ss_compounds):
        errors = []
        warnings = []
        if not blanks:
            errors.append('未识别到 BLANK 列；请检查列名是否包含 BLANK。')
        if not samps:
            errors.append('未识别到 sample 列；请检查样品列名称。')
        if not target_compounds:
            errors.append('未识别到目标化合物行；请检查化合物名称列。')
        if not mss:
            warnings.append('未识别到 MS/基质加标列；基质加标回收率将无法计算。')
        if len(blanks) < 2:
            warnings.append('BLANK 列少于 2 个；请确认 MDL 计算所需的空白数量。')
        if not is_compounds:
            warnings.append('未识别到 IS 内标；请确认是否使用内标校正。')
        if not ss_compounds:
            warnings.append('未识别到 SS 替代物；SS 回收率不会自动生成。')
        return {
            'ready': not errors,
            'errors': errors,
            'warnings': warnings,
            'summary': f'{len(blanks)} BLANK + {len(mss)} MS + {len(samps)} sample + '
                       f'{len(is_compounds)} IS + {len(ss_compounds)} SS + {len(target_compounds)} 个目标物',
        }


def read_preview_table(file_bytes):
    """Read XLSX/XLS or a delimited MassHunter export for display."""
    if file_bytes.startswith(b'PK'):
        return pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, header=None)
    if file_bytes.startswith(b'\xd0\xcf\x11\xe0'):
        return pd.read_excel(io.BytesIO(file_bytes), header=None, engine='xlrd')
    for encoding in ('utf-8-sig', 'utf-16', 'utf-8', 'gb18030', 'big5'):
        try:
            return pd.read_csv(io.BytesIO(file_bytes), header=None, encoding=encoding, sep=None, engine='python')
        except (UnicodeDecodeError, pd.errors.ParserError):
            continue
    raise ValueError('无法识别 CSV 编码或分隔符，请保存为 UTF-8、UTF-16、GB18030 或制表符分隔的 CSV。')

# ============================================================
# 中英文对照字典
# ============================================================
T = {
    'page_title':           {'zh': 'LC-MS/MS 数据处理智能体',                    'en': 'LC-MS/MS Data Processing Agent'},
    'page_subtitle':        {'zh': '上传 MassHunter 原始数据 → 一键生成完整分析表格 → 下载', 'en': 'Upload MassHunter raw data → Auto-generate standard tables → Download'},
    'sidebar_header':       {'zh': '📋 实验参数',                                'en': '📋 Experiment Parameters'},
    'sample_type':          {'zh': '样本类型',                                   'en': 'Sample Type'},
    'sample_type_opts':     {'zh': ['尿液','灰尘','土壤','水','其他'],             'en': ['Urine','Dust','Soil','Water','Other']},
    'sample_vol':           {'zh': '取样量 (mL/g)',                              'en': 'Sample Volume (mL/g)'},
    'final_vol':            {'zh': '定容体积 (mL)',                              'en': 'Final Volume (mL)'},
    'extra_dil':            {'zh': '额外稀释倍数',                                'en': 'Extra Dilution Factor'},
    'is_correction':        {'zh': '数据是否经过内标（IS）校正？',                 'en': 'IS (Internal Standard) Correction Applied?'},
    'is_opts':              {'zh': ['✅ 是，有内标校正','❌ 否，无内标校正'],       'en': ['✅ Yes, IS Corrected','❌ No, Raw Concentration']},
    'is_help':              {'zh': '内标校正后，导出值已是原始样本浓度；无内标校正则需要手动换算',
                             'en': 'With IS correction, exported values are already original sample concentration; without IS correction, manual conversion is needed'},
    'csv_note':             {'zh': 'CSV 是单页、无公式数值报告；需要多工作表和可审计公式时请选择 XLSX。', 'en': 'CSV is a flat, formula-free report; choose XLSX for multi-sheet formulas.'},
    'cf_caption_yes':       {'zh': '💡 有内标校正，仪器已自动将进样瓶浓度换算为原始尿液浓度，换算因子 = 1',
                             'en': '💡 IS corrected: instrument has already converted vial concentration to original sample concentration. CF = 1'},
    'cf_caption_no':        {'zh': '💡 无内标校正，导出值是进样瓶浓度，换算因子用于将其换算为原始尿液浓度：定容体积 ÷ 取样量 × 稀释倍数',
                             'en': '💡 Not IS corrected: exported value is vial concentration. CF = Final Vol ÷ Sample Vol × Dilution to convert to original concentration'},
    'cf_label':             {'zh': '换算因子',                                   'en': 'Conversion Factor'},
    'cf_locked':            {'zh': '（已锁定）',                                  'en': ' (Locked)'},
    'cf_editable':          {'zh': '（可手动覆盖）',                              'en': ' (Editable)'},
    'spike_conc':           {'zh': '基质加标浓度 (ppb)',                          'en': 'Matrix Spike Conc (ppb)'},
    'file_header':          {'zh': '📁 文件',                                    'en': '📁 File'},
    'upload_label':         {'zh': '上传原始数据（XLSX 或 CSV）',                  'en': 'Upload Raw Data (XLSX or CSV)'},
    'output_name':          {'zh': '输出文件名',                                  'en': 'Output Filename'},
    'output_default':       {'zh': '已处理数据.xlsx',                             'en': 'processed_data.xlsx'},
    'output_format':        {'zh': '输出格式',                                    'en': 'Output Format'},
    'roles_header':         {'zh': '化合物角色设置',                              'en': 'Compound Roles'},
    'roles_caption':        {'zh': '系统按名称预识别；请确认哪些为 IS、SS。未选为 IS/SS 的化合物将作为目标物。', 'en': 'Roles are auto-detected by name; confirm IS and SS. Unselected compounds remain targets.'},
    'is_select':            {'zh': 'IS 内标（可多选）',                            'en': 'IS internal standards'},
    'ss_select':            {'zh': 'SS 替代物（可多选）',                          'en': 'SS surrogates'},
    'blank_zero_header':    {'zh': 'blank=0 的 MDL 设置',                        'en': 'Blank-zero MDL settings'},
    'ss_spike_grid':        {'zh': '已选 SS 的理论加标浓度（ppb）',                   'en': 'Theoretical spike concentration for selected SS (ppb)'},
    'custom_ss':            {'zh': '自定义 SS 替代物（可选）',                       'en': 'Custom SS surrogates (optional)'},
    'custom_ss_help':       {'zh': '每行输入“名称, 理论加标浓度(ppb)”。示例：d7-C12-BAC, 4。名称必须与上传文件的化合物名称一致；系统将其列为 SS，并按“SS 实测值 ÷ 该 SS 加标浓度 × 100%”计算。', 'en': 'One per line: “name, theoretical spike concentration (ppb)”. Example: d7-C12-BAC, 4. The name must match an imported compound; it will be treated as SS and recovery = measured SS ÷ its spike concentration × 100%.'},
    'custom_ss_placeholder': {'zh': 'd7-C12-BAC, 4\nMy Surrogate, 2.5',             'en': 'd7-C12-BAC, 4\nMy Surrogate, 2.5'},
    'blank_zero_select':    {'zh': '选择 blank=0 的化合物',                         'en': 'Select blank=0 compounds'},
    'blank_zero_help':      {'zh': '对每个所选化合物输入标曲浓度和对应 S/N：MDL = 3 × 标曲浓度 ÷ S/N。', 'en': 'Enter calibration concentration and S/N for each selected compound: MDL = 3 × calibration concentration ÷ S/N.'},
    'calibration':          {'zh': '标曲浓度 (ppb)',                              'en': 'Calibration concentration (ppb)'},
    'sn':                   {'zh': 'S/N',                                         'en': 'S/N'},
    'mql_help':             {'zh': '默认 3.333333；请按实验室方法确认。',              'en': 'Default 3.333333; confirm with your laboratory method.'},
    'preview_stats':        {'zh': '描述性统计（在线数值预览）',                     'en': 'Descriptive statistics (numeric preview)'},
    'preview_stats_help':   {'zh': '这里显示已计算的数值；下载的 Excel 同时保留可审计公式。', 'en': 'Calculated values are shown here; downloaded Excel retains auditable formulas.'},
    'preview_final':        {'zh': '最终浓度（在线数值预览）',                       'en': 'Final concentrations (numeric preview)'},
    'tip':                  {'zh': '上传文件 → 调参数 → 点处理 → 下载结果',        'en': 'Upload → Adjust Params → Process → Download'},
    'demo_btn':             {'zh': '📥 加载 Demo 数据',                           'en': '📥 Load Demo Data'},
    'demo_help':            {'zh': '使用示例数据体验平台功能，无需上传自己的文件',    'en': 'Try the platform with example data, no upload needed'},
    'raw_preview':          {'zh': '📊 原始数据预览',                             'en': '📊 Raw Data Preview'},
    'total_cmpd':           {'zh': '化合物总数',                                  'en': 'Total Compounds'},
    'target_cmpd':          {'zh': '目标化合物',                                  'en': 'Target Compounds'},
    'is_cmpd':              {'zh': 'IS内标',                                     'en': 'IS Internal Std'},
    'ss_cmpd':              {'zh': 'SS替代物',                                    'en': 'SS Surrogates'},
    'view_classify':        {'zh': '查看化合物分类',                              'en': 'View Compound Classification'},
    'target_label':         {'zh': '目标化合物',                                  'en': 'Target Compounds'},
    'is_label':             {'zh': 'IS内标',                                     'en': 'IS Internal Standards'},
    'ss_label':             {'zh': 'SS替代物',                                    'en': 'SS Surrogates'},
    'blank_cols':           {'zh': 'BLANK列',                                    'en': 'BLANK Columns'},
    'ms_cols':              {'zh': 'MS列',                                       'en': 'MS Columns'},
    'sample_cols':          {'zh': '样品列',                                      'en': 'Sample Columns'},
    'preview_warn':         {'zh': '预览时请注意',                                 'en': 'Preview Warning'},
    'process_btn':          {'zh': '🚀 开始处理',                                 'en': '🚀 Start Processing'},
    'processing':           {'zh': '正在处理数据...',                              'en': 'Processing data...'},
    'success':              {'zh': '✅ 处理完成！',                                'en': '✅ Processing Complete!'},
    'result_preview':       {'zh': '📋 处理结果预览',                             'en': '📋 Result Preview'},
    'rows_cols':            {'zh': '行 × ',                                      'en': ' rows × '},
    'rows_cols_suffix':     {'zh': '列',                                         'en': ' columns'},
    'download_btn':         {'zh': '⬇️ 下载处理结果',                             'en': '⬇️ Download Result'},
    'error':                {'zh': '处理失败',                                    'en': 'Processing Failed'},
}

def t(key, lang='zh'):
    """Get translation for key in given language"""
    if key in T:
        return T[key].get(lang, T[key].get('zh', key))
    return key

# ============================================================
# 页面设置
# ============================================================
st.set_page_config(
    page_title="LC-MS/MS 数据处理智能体 | LC-MS/MS Data Processor",
    page_icon="🔬",
    layout="wide",
)

# 语言切换
with st.sidebar:
    lang = st.radio("🌐 Language / 语言", ["中文", "English"], index=0, horizontal=True, label_visibility="collapsed")
    L = 'zh' if lang == '中文' else 'en'

st.title(f"🔬 {t('page_title', L)}")
st.markdown(t('page_subtitle', L))

# ============================================================
# 侧边栏
# ============================================================
with st.sidebar:
    st.header(t('sidebar_header', L))

    sample_type = st.selectbox(
        t('sample_type', L),
        t('sample_type_opts', L),
        index=0
    )
    # Map back to Chinese for internal use
    zh_sample_type = {'Urine':'尿液','Dust':'灰尘','Soil':'土壤','Water':'水','Other':'其他'}.get(sample_type, sample_type)
    if L == 'zh': zh_sample_type = sample_type
    else: sample_type = zh_sample_type

    output_unit = "ng/mL" if sample_type == "尿液" else ("ng/g" if sample_type in ("灰尘","土壤") else "ng/mL")

    col1, col2 = st.columns(2)
    with col1:
        sample_vol = st.number_input(t('sample_vol', L), value=2.0, step=0.1, format="%.1f")
    with col2:
        final_vol = st.number_input(t('final_vol', L), value=0.5, step=0.1, format="%.1f")

    extra_dil = st.number_input(t('extra_dil', L), value=1, step=1, min_value=1)

    use_is = st.radio(
        t('is_correction', L),
        options=t('is_opts', L),
        index=0,
        help=t('is_help', L)
    )
    is_corrected = use_is.startswith("✅")

    if is_corrected:
        auto_cf = 1.0
        st.caption(t('cf_caption_yes', L))
    else:
        auto_cf = round(final_vol / sample_vol * extra_dil, 6)
        st.caption(t('cf_caption_no', L))

    cf_locked_text = t('cf_locked', L) if is_corrected else t('cf_editable', L)
    conversion_factor = st.number_input(
        t('cf_label', L) + cf_locked_text,
        value=auto_cf,
        step=0.001,
        format="%.6f",
        disabled=is_corrected
    )
    spike_conc = st.number_input(t('spike_conc', L), value=10, step=1)

    st.subheader(t('custom_ss', L))
    st.caption(t('custom_ss_help', L))
    custom_ss_text = st.text_area(
        t('custom_ss', L),
        value='',
        placeholder=t('custom_ss_placeholder', L),
        key='custom_ss_text',
        label_visibility='collapsed',
    )

    st.divider()
    st.header(t('file_header', L))
    uploaded_file = st.file_uploader(t('upload_label', L), type=["xlsx", "xls", "csv", "tsv"])

    output_name = st.text_input(t('output_name', L), t('output_default', L))
    output_format = st.selectbox(t('output_format', L), ['XLSX', 'CSV'])
    st.caption(t('csv_note', L))

    st.divider()
    st.caption(t('tip', L))

# ============================================================
# 主区域
# ============================================================

def get_demo_bytes():
    """生成 Demo 示例数据（内置，无需外部文件）"""
    import openpyxl as _xl
    wb = _xl.Workbook()
    ws = wb.active; ws.title = 'Sheet1'
    ws.cell(row=1, column=1, value='化合物方法')
    ws.cell(row=2, column=1, value='名称'); ws.cell(row=2, column=2, value='离子对')
    hdrs = ['F91-BLANK1','F92-BLANK2','F93-BLANK3','F94-BLANK4','F95-BLANK5','F96-BLANK6',
            '10PPB','F89-MS1','F90-MS2','F1','F2','F3','F4','F5','F6','F7','F8','F9','F10']
    for i,h in enumerate(hdrs):
        ws.cell(row=1, column=3+i, value=h); ws.cell(row=2, column=3+i, value='最终浓度')
    demo_data = [
        ('C8-BAC','248.2->91.0', [0.000498,0.019774,0.000505,0.000671,0.000549,0.001351], [0.1741,9.9987], [0.2617,0.2135], [0.000919,0.000893,0.002015,None,0.000692,0.001064,0.000784,0.001488,0.000453,None]),
        ('C10-BAC','276.3->91.1', [0.001464,0.023144,0.001582,0.003866,0.001540,0.010050], [0.1505,10.0234], [1.3550,0.9567], [0.0257,0.0261,0.0247,None,0.0243,0.0109,0.0200,0.0184,0.0179,0.0057]),
        ('C12-BAC','304.3->91.0', [0.062243,0.068513,0.053031,0.059037,0.057253,0.152859], [0.4065,10.1234], [1.2908,0.9064], [0.0875,0.0823,0.0912,None,0.0789,0.0680,0.0734,0.0710,0.0663,0.0758]),
        ('C14-BAC','332.3->91.1', [0.012054,0.020653,0.010276,0.011484,0.010341,0.032861], [0.1965,10.0345], [0.4191,0.3179], [0.0268,0.0241,0.0289,None,0.0217,0.0193,0.0224,0.0208,0.0185,0.0246]),
        ('C16-BAC','360.4->91.1', [0.009222,0.019469,0.009451,0.007037,0.009124,0.012168], [0.1643,10.0456], [0.1224,0.1248], [0.0156,0.0139,0.0168,None,0.0121,0.0105,0.0118,0.0112,0.0098,0.0134]),
        ('C18-BAC','388.4->91.0', [0.017903,0.025084,0.017392,0.015902,0.016529,0.035585], [0.1479,10.0678], [0.0635,0.0646], [0.0289,0.0256,0.0312,None,0.0224,0.0198,0.0220,0.0205,0.0183,0.0248]),
        ('C8-DDAC','270.3->158.2', [0.012746,0.012324,0.010552,0.011981,0.012511,0.015686], [1.9238,10.0890], [1.9238,1.7021], [0.0198,0.0176,0.0215,None,0.0154,0.0137,0.0152,0.0141,0.0126,0.0172]),
        ('C12-DDAC','382.4->214.0', [0.009503,0.012877,0.009371,0.008498,0.009007,0.016294], [0.4065,10.1112], [0.4065,0.5414], [0.0152,0.0135,0.0165,None,0.0118,0.0104,0.0116,0.0108,0.0096,0.0132]),
        ('C8-ATMAC','172.2->71.1', [0.013368,0.015907,0.009256,0.011528,0.009870,0.011313], [1.5743,10.1334], [1.5743,1.1097], [0.0168,0.0149,0.0182,None,0.0131,0.0115,0.0128,0.0119,0.0106,0.0146]),
        ('C12-ATMAC','228.3->71.1', [0.161189,0.250358,0.186107,0.199182,0.182527,0.264367], [1.1791,10.1556], [1.1791,0.9213], [0.2567,0.2289,0.2789,None,0.1998,0.1765,0.1956,0.1823,0.1623,0.2234]),
        ('C8-PFAS','499.0->80.0', [0.001,0.002,0.001,0.002,0.001,0.002], [0.0,10.0], [0.0,0.0], [0.02,0.03,None,0.01,0.04,0.02,0.01,0.03,0.02,0.01]),
        ('C10-BAC+O','292.3->91.1', [0.000629,0.001272,0.000783,0.001354,0.000741,0.003236], [2.0656,10.1778], [1.8895,1.3102], [0.0021,0.0019,0.0023,None,0.0017,0.0015,0.0016,0.0015,0.0013,0.0018]),
        ('d7-C12-BAC','311.3->98.1', [0.4801,0.4331,0.5060,0.5890,0.4644,1.0049], [0.275,9.988], [1.1597,0.7278], []),
        ('d9-C10-ATMAC','209.3->71.1', [0.7596,1.5442,1.0103,0.9204,1.1522,2.0779], [0.200,10.001], [1.2243,0.8579], []),
    ]
    for row_idx, (name, ion, blanks, qcs, mss, samps) in enumerate(demo_data):
        r = 3 + row_idx
        ws.cell(row=r, column=1, value=name); ws.cell(row=r, column=2, value=ion)
        for j, v in enumerate(blanks):
            if v is not None: ws.cell(row=r, column=3+j, value=round(v, 6))
        for j, v in enumerate(mss):
            if v is not None: ws.cell(row=r, column=10+j, value=round(v, 6))
        for j, v in enumerate(samps):
            if v is not None: ws.cell(row=r, column=12+j, value=round(v, 6))
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

demo_path = os.path.join(os.path.dirname(__file__), 'demo_urine_qac_masshunter.xlsx')
with st.sidebar:
    use_demo = st.button(t('demo_btn', L), help=t('demo_help', L), use_container_width=True)

if use_demo:
    if os.path.exists(demo_path):
        with open(demo_path, 'rb') as demo_file:
            st.session_state.demo_bytes = demo_file.read()
        st.session_state.demo_active = True
    else:
        st.error('Demo file not found. Please upload your own file.')

file_bytes = st.session_state.get('demo_bytes') if st.session_state.get('demo_active') else None
if uploaded_file:
    file_bytes = uploaded_file.getvalue()
    st.session_state.demo_active = False
    st.session_state.pop('demo_bytes', None)

if st.session_state.get('demo_active') and file_bytes:
    st.info('Demo 数据已加载：这是可直接处理的尿液 MassHunter 示例；可检查识别结果后点击“开始处理”。')

selected_is = []
selected_ss = []
ss_concentrations = {}
mdl_overrides = {}
mql_factor = 3.333333
layout_is_ready = False

if file_bytes:
    st.subheader(t('raw_preview', L))
    try:
        raw_data, blanks, mss, samps, target, is_c, ss_c, all_c = read_raw(file_bytes)

        layout_report = validate_input_layout(blanks, mss, samps, target, is_c, ss_c)
        layout_is_ready = layout_report['ready']
        if layout_report['ready']:
            st.success(f"文件格式检查通过：{layout_report['summary']}")
        else:
            st.error('文件格式检查未通过：' + '；'.join(layout_report['errors']))
        for message in layout_report['warnings']:
            st.warning(message)

        df_raw = read_preview_table(file_bytes)
        st.dataframe(df_raw.head(8), use_container_width=True)

        col1, col2, col3, col4 = st.columns(4)
        col1.metric(t('total_cmpd', L), len(all_c))
        col2.metric(t('target_cmpd', L), len(target))
        col3.metric(t('is_cmpd', L), len(is_c))
        col4.metric(t('ss_cmpd', L), len(ss_c))

        col1, col2, col3 = st.columns(3)
        col1.metric(t('blank_cols', L), len(blanks))
        col2.metric(t('ms_cols', L), len(mss))
        col3.metric(t('sample_cols', L), len(samps))

        st.subheader(t('roles_header', L))
        st.caption(t('roles_caption', L))
        selected_is = st.multiselect(t('is_select', L), all_c, default=is_c)
        selected_ss = st.multiselect(t('ss_select', L), all_c, default=ss_c)
        custom_ss, custom_ss_errors = parse_custom_ss_entries(custom_ss_text)
        for message in custom_ss_errors:
            st.error(message)
        missing_custom_ss = [name for name in custom_ss if name not in all_c]
        if missing_custom_ss:
            st.error('未在上传文件中找到自定义 SS：' + ', '.join(missing_custom_ss))
        valid_custom_ss = [name for name in custom_ss if name in all_c]
        selected_ss = list(dict.fromkeys(selected_ss + valid_custom_ss))
        overlap = set(selected_is) & set(selected_ss)
        if overlap:
            st.error(f'同一化合物不能同时作为 IS 与 SS：{sorted(overlap)}')
        if selected_ss:
            st.write(t('ss_spike_grid', L))
            ss_grid = st.columns(min(3, len(selected_ss)))
            for i, name in enumerate(selected_ss):
                with ss_grid[i % len(ss_grid)]:
                    if name in custom_ss:
                        # A custom SS line is the source of truth for its own
                        # spike concentration.  Show the exact calculation
                        # value read-only so displayed and used values match.
                        ss_concentrations[name] = st.number_input(
                            name, min_value=0.000001, value=float(custom_ss.get(name, 4.0)),
                            step=1.0, key=f'custom_ss_conc_{name}', disabled=True,
                        )
                    else:
                        ss_concentrations[name] = st.number_input(
                            name, min_value=0.000001, value=4.0, step=1.0, key=f'ss_conc_{name}'
                        )

        # This table reflects the final user-confirmed IS/SS selection, not
        # merely the name-pattern auto-detection shown immediately on import.
        with st.expander(t('view_classify', L)):
            roles_for_display = resolve_roles(all_c, selected_is, selected_ss)
            st.write(f"**{t('target_label', L)}**:", ", ".join(roles_for_display['target_compounds']) if roles_for_display['target_compounds'] else "-")
            st.write(f"**{t('is_label', L)}**:", ", ".join(roles_for_display['is_compounds']) if roles_for_display['is_compounds'] else "-")
            st.write(f"**{t('ss_label', L)}**:", ", ".join(roles_for_display['ss_compounds']) if roles_for_display['ss_compounds'] else "-")
            classification_rows = compound_classification_rows(all_c, selected_is, selected_ss)
            st.dataframe(pd.DataFrame(classification_rows), use_container_width=True, hide_index=True)

        with st.expander(t('blank_zero_header', L)):
            detected_blank_zero = detect_blank_zero_compounds(raw_data, blanks)
            blank_zero_compounds = st.multiselect(
                t('blank_zero_select', L), all_c,
                default=[name for name in detected_blank_zero if name in all_c],
            )
            if blank_zero_compounds:
                st.caption(t('blank_zero_help', L))
                mdl_cols = st.columns(min(3, len(blank_zero_compounds)))
                for i, name in enumerate(blank_zero_compounds):
                    with mdl_cols[i % len(mdl_cols)]:
                        calibration = st.number_input(
                            f'{name} {t("calibration", L)}', min_value=0.0, value=0.0,
                            step=0.1, key=f'mdl_cal_{name}'
                        )
                        sn = st.number_input(
                            f'{name} {t("sn", L)}', min_value=0.0, value=0.0,
                            step=1.0, key=f'mdl_sn_{name}'
                        )
                        mdl_overrides[name] = {
                            'blank_zero': True,
                            'calibration_concentration': calibration,
                            'signal_to_noise': sn,
                        }

        mql_factor = st.number_input(
            'MQL / MDL 倍数', min_value=0.000001, value=3.333333,
            step=0.1, format='%.6f', help=t('mql_help', L)
        )

    except Exception as e:
        st.warning(f"{t('preview_warn', L)}: {e}")

# ============================================================
# 处理按钮
# ============================================================
st.divider()
process_btn = st.button(t('process_btn', L), type="primary", disabled=(file_bytes is None or not layout_is_ready), use_container_width=True)

if process_btn and file_bytes:
    with st.spinner(t('processing', L)):

        config = {
            'sample_type': sample_type,
            'sample_volume_ml': float(sample_vol),
            'final_volume_ml': float(final_vol),
            'extra_dilution': int(extra_dil),
            'conversion_factor': float(conversion_factor),
            'spike_conc_ppb': int(spike_conc),
            'is_compounds': selected_is,
            'ss_compounds': selected_ss,
            'ss_spike_concentrations': ss_concentrations,
            'mdl_overrides': mdl_overrides,
            'mql_factor': float(mql_factor),
            'masshunter_unit': 'ppb',
            'output_unit': output_unit,
            'blank_handling': 'ND',
            'input_file': '',
            'output_file': output_name,
            'output_format': output_format.lower(),
            'input_bytes': file_bytes,
        }

        try:
            output_bytes, filename = process(config=config, return_bytes=True)

            st.success(t('success', L))

            roles = resolve_roles(all_c, selected_is, selected_ss)
            preview_cfg = {**config, 'target_compounds': roles['target_compounds']}
            preview_rows = compute_preview_summary(raw_data, blanks, samps, preview_cfg)
            if preview_rows:
                st.subheader(t('preview_stats', L))
                st.caption(t('preview_stats_help', L))
                st.dataframe(pd.DataFrame(preview_rows), use_container_width=True)

            final_preview_rows = compute_preview_final_table(raw_data, blanks, samps, preview_cfg)
            if final_preview_rows:
                st.subheader(t('preview_final', L))
                st.dataframe(pd.DataFrame(final_preview_rows), use_container_width=True)

            st.subheader(t('result_preview', L))
            if output_format == 'CSV':
                # CSV is a flat, formula-free report; preview the actual exported values.
                # The summary and final-concentration sections have different widths,
                # so read rows directly and pad them before displaying the table.
                import csv
                csv_rows = list(csv.reader(io.StringIO(output_bytes.decode('utf-8-sig'))))
                width = max((len(row) for row in csv_rows), default=0)
                csv_preview = pd.DataFrame([row + [''] * (width - len(row)) for row in csv_rows])
                st.dataframe(csv_preview, use_container_width=True)
            else:
                # openpyxl cannot calculate formulas. Show the numerically evaluated
                # previews above instead of exposing formula strings in the result view.
                st.info(t('preview_stats_help', L))
                st.caption('Excel 下载文件保留可审计公式；在线预览显示按同一规则计算的数值。')
                st.dataframe(pd.DataFrame(preview_rows), use_container_width=True)
                st.dataframe(pd.DataFrame(final_preview_rows), use_container_width=True)

            st.divider()
            st.download_button(
                label=t('download_btn', L),
                data=output_bytes,
                file_name=filename,
                mime=("text/csv" if output_format == 'CSV' else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                use_container_width=True,
                type="primary",
            )

        except Exception as e:
            st.error(f"{t('error', L)}: {e}")
            import traceback
            st.code(traceback.format_exc())
