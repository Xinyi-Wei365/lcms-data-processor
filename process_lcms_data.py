#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
LC-MS/MS 数据处理智能体 v2
从 MassHunter 导出的原始数据生成最终计算浓度表格

用法：
  python process_lcms_data.py              # 使用 CONFIG 中的参数直接运行
  from process_lcms_data import process     # 作为模块导入，供 Streamlit 调用
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import statistics
import math
import re
import io
import csv
import os
import codecs
import zipfile
import xml.etree.ElementTree as ET
import pandas as pd


# One-sided 99 % Student-t critical values.  The values are used only for the
# numeric web/CSV preview; XLSX uses legacy-compatible TINV/STDEV formulas so
# both Microsoft Excel and WPS can recalculate without #NAME? errors.
T99_ONE_SIDED = {
    1: 31.821, 2: 6.965, 3: 4.541, 4: 3.747, 5: 3.365, 6: 3.143,
    7: 2.998, 8: 2.896, 9: 2.821, 10: 2.764, 11: 2.718, 12: 2.681,
    13: 2.650, 14: 2.624, 15: 2.602, 16: 2.583, 17: 2.567, 18: 2.552,
    19: 2.539, 20: 2.528, 21: 2.518, 22: 2.508, 23: 2.500, 24: 2.492,
    25: 2.485, 26: 2.479, 27: 2.473, 28: 2.467, 29: 2.462, 30: 2.457,
}


WORKBOOK_TEXT = {
    'zh': {
        'matrix': 'Matrix spike  基质加标浓度', 'blank': 'Blanks_MDL 空白基质检出限',
        'bottle': 'Conc. in bottle 瓶内实测浓度', 'final': 'Final. conc 最终计算浓度',
        'summary': '描述性统计', 'notes': '计算说明',
        'compound': '化合物名称', 'group': '分组', 'sample_volume': '取样体积 (mL)',
        'is_additions': 'IS measured concentrations, ppb  IS实测浓度（来源于原始MS列）',
        'ss_recoveries': 'SS recoveries, %  替代物回收率',
        'summary_title': '描述性统计（最终结果，保留3位有效数字）',
        'summary_note': 'DF=有数值最终浓度数÷全部样品列数；仅DF>50%时展示Median(Q1-Q3)；MDL/MQL保留公式。',
        'summary_headers': ['名称', '链长', 'DF (%)', 'Median (Q1-Q3)', 'MDL', 'MQL'],
        'blank_note': '数值空白单元格不参与统计；有效但未检出样品按 1/2 MDL 替代。',
        'notes_title': 'LC-MS/MS 数据处理说明',
    },
    'en': {
        'matrix': 'Matrix spike', 'blank': 'Blanks_MDL', 'bottle': 'Concentration in vial',
        'final': 'Final concentration', 'summary': 'Descriptive statistics', 'notes': 'Calculation notes',
        'compound': 'Compound name', 'group': 'Group', 'sample_volume': 'Sample volume (mL)',
        'is_additions': 'IS measured concentrations, ppb (from raw MS columns)', 'ss_recoveries': 'SS recoveries, %',
        'summary_title': 'Descriptive statistics (final results, 3 significant figures)',
        'summary_note': 'DF=numeric final-result cells÷all sample columns; Median(Q1-Q3) is shown only when DF>50%; MDL/MQL formulas are retained.',
        'summary_headers': ['Name', 'Chain length', 'DF (%)', 'Median (Q1-Q3)', 'MDL', 'MQL'],
        'blank_note': 'Blank cells are excluded; valid non-detects are replaced with 1/2 MDL.',
        'notes_title': 'LC-MS/MS calculation notes',
    },
}


def workbook_text(cfg):
    return WORKBOOK_TEXT['en' if (cfg or {}).get('language') == 'en' else 'zh']


def sheet_name(cfg, key):
    return workbook_text(cfg)[key]

# ============================================================
# 可配置参数
# ============================================================
CONFIG = {
    'sample_type': '尿液',
    'sample_volume_ml': 2,
    'final_volume_ml': 0.5,
    'extra_dilution': 1,
    'conversion_factor': 1,
    'spike_conc_ppb': 10,
    'masshunter_unit': 'ppb',
    'output_unit': 'ng/mL',
    'blank_handling': 'ND',
    'input_file': r'C:\Users\HP-PC\Desktop\未整理 - 副本.xlsx',
    'output_file': r'C:\Users\HP-PC\Desktop\已处理_尿液数据.xlsx',
}

# ============================================================
# 样式常量
# ============================================================
FONT_COLOR = '384350'
HEADER_BG = 'E1E3E5'
BLUE_FONT = '1F4E78'
BLUE_BG = 'D9EAF7'
GOLD_FONT = '7F6000'
GOLD_BG = 'FFF2CC'
RED_FONT = 'C00000'
YELLOW_BG = 'FFFF00'

# ============================================================
# 化合物分组（按输出顺序）
# ============================================================
BAC_SERIES   = ['C8-BAC','C10-BAC','C12-BAC','C14-BAC','C16-BAC','C18-BAC']
DDAC_SERIES  = ['C8-DADMAC','C8-10-DADMAC','C10-DADMAC','C12-DADMAC','C14-DADMAC','C16-DADMAC','C18-DADMAC']
ATMAC_SERIES = ['C8-ATMAC','C10-ATMAC','C12-ATMAC','C14-ATMAC','C16-ATMAC','C18-ATMAC']
BAC_COOH     = ['C10-BAC+2O-2H','C12-BAC+2O-2H','C14-BAC+2O-2H']
BAC_OH       = ['C10-BAC+O','C12-BAC+O','C14-BAC+O']
IS_COMPS     = ['C10-BAC-COOH[C13]','C12-BAC+COOH-d6','C12-BAC+OH-d3']
SS_COMPS     = ['d7-C12-BAC','d9-C10-ATMAC']

TARGET_COMPS = BAC_SERIES + DDAC_SERIES + ATMAC_SERIES + BAC_COOH + BAC_OH
ALL_COMPS    = TARGET_COMPS + IS_COMPS + SS_COMPS


# ============================================================
# 工具函数
# ============================================================
def safe_float(v):
    if v is None: return None
    try: return float(v)
    except: return None


def is_empty_cell(value):
    """Return True only for genuinely empty source cells, not ND or other text."""
    return value is None or (isinstance(value, str) and not value.strip())


def normalize_analyte_name(name):
    """Normalize the requested DDAC display spelling to DADMAC."""
    original = str(name or '').strip()
    return re.sub(r'(?i)(?<![A-Z])DDAC(?![A-Z])', 'DADMAC', original)

def round6(v):
    if v is None: return None
    return round(float(v), 6)

def round_int(v):
    if v is None: return None
    return round(float(v))


def extract_chain_length(name):
    """Return the first C-number chain label without changing the analyte name."""
    match = re.search(r'(?<![A-Za-z])C(\d+)(?!\d)', str(name or ''), flags=re.I)
    return f'C{match.group(1)}' if match else None


def infer_analyte_type(name):
    """Infer a sortable chemical family from the imported analyte name.

    QAC family names keep their established labels.  For other chemicals with
    a chain-length prefix (for example ``C8-PFAS`` or ``C10-Phthalate``), the
    text after the chain length becomes a family label.  Unknown free-text
    names remain ``Other`` rather than being silently treated as QACs.
    """
    upper = normalize_analyte_name(name).upper()
    if 'ATMAC' in upper:
        return 'ATMAC'
    if 'DADMAC' in upper:
        return 'DADMAC'
    if 'BAC' in upper:
        return 'BAC'

    # Prefer well-known multi-character families when present anywhere in a
    # method name, including names without an explicit C-number prefix.
    for family in ('PFOS', 'PFOA', 'PFHxS', 'PFNA', 'PFDA', 'PFAS', 'PAH', 'PCB'):
        if family.upper() in upper:
            return family.upper()

    # Generic non-QAC family: capture the chemical-type part after C<number>.
    # Remove isotope/oxidation decorations so related analytes stay together.
    match = re.search(r'(?:^|[-_\s])C\d+(?:-\d+)?[-_\s]+([A-Za-z][A-Za-z0-9 ]*)', upper)
    if match:
        family = re.split(r'\s*(?:\+|-)(?:\d*O|\d*H|D\d+|\[.*)', match.group(1), maxsplit=1)[0]
        family = re.sub(r'\s+', ' ', family).strip(' -_')
        if family and family not in {'OTHER', 'UNKNOWN'}:
            return family
    return 'Other'


def analyte_metadata(name):
    """Return display-safe metadata while preserving the original analyte name."""
    original = normalize_analyte_name(name)
    return {
        'name': original,
        'type': infer_analyte_type(original),
        'chain_length': extract_chain_length(original),
    }


def compound_metadata_for(name, overrides=None):
    """Return name metadata after applying a user-confirmed classification."""
    normalized = normalize_analyte_name(name)
    automatic = analyte_metadata(normalized)
    override = (overrides or {}).get(normalized) or (overrides or {}).get(name) or {}
    role = str(override.get('role') or '').strip()
    chain = str(override.get('chain_length') or automatic['chain_length'] or 'NA').strip()
    if chain.upper() != 'NA' and re.fullmatch(r'\d+', chain):
        chain = f'C{chain}'
    return {
        'name': normalized,
        'type': str(override.get('type') or automatic['type'] or 'Other').strip() or 'Other',
        'chain_length': chain or 'NA',
        'role': role,
    }


def normalize_compound_role(role):
    """Normalize bilingual UI role labels to the three processing roles."""
    value = str(role or '').strip().lower()
    if value in {'is', 'internal standard', '内标'}:
        return 'IS'
    if value in {'ss', 'surrogate', '替代物'}:
        return 'SS'
    if value in {'target', 'target analyte', '目标物'}:
        return 'Target'
    return ''


def resolve_roles(compounds, is_compounds=None, ss_compounds=None):
    """Resolve roles from the detected list; unselected compounds remain targets."""
    all_compounds = list(dict.fromkeys(normalize_analyte_name(value) for value in compounds if str(value).strip()))
    is_set = {normalize_analyte_name(value) for value in (is_compounds or [])}
    ss_set = {normalize_analyte_name(value) for value in (ss_compounds or [])}
    overlap = is_set & ss_set
    if overlap:
        raise ValueError(f'An analyte cannot be both IS and SS: {sorted(overlap)}')
    is_list = [name for name in all_compounds if name in is_set]
    ss_list = [name for name in all_compounds if name in ss_set]
    targets = [name for name in all_compounds if name not in is_set and name not in ss_set]
    return {'target_compounds': targets, 'is_compounds': is_list, 'ss_compounds': ss_list}


def configured_compound_lists(cfg, detected_compounds, detected_is=None, detected_ss=None):
    """Return target/IS/SS lists using user selections with detected-role defaults."""
    overrides = cfg.get('compound_metadata') or {}
    explicit_is = list(cfg.get('is_compounds', detected_is or []))
    explicit_ss = list(cfg.get('ss_compounds', detected_ss or []))
    for name in detected_compounds:
        role = normalize_compound_role(compound_metadata_for(name, overrides).get('role'))
        if role == 'IS' and name not in explicit_is:
            explicit_is.append(name)
        elif role == 'SS' and name not in explicit_ss:
            explicit_ss.append(name)
    roles = resolve_roles(
        detected_compounds,
        explicit_is,
        explicit_ss,
    )
    ordered_targets = sort_compounds(roles['target_compounds'], overrides)
    ordered_is = sort_compounds(roles['is_compounds'], overrides)
    ordered_ss = sort_compounds(roles['ss_compounds'], overrides)
    return ordered_targets, ordered_is, ordered_ss, ordered_targets + ordered_is + ordered_ss


def compound_classification_rows(compounds, is_compounds=None, ss_compounds=None, overrides=None):
    """Create an inspectable classification table for the Streamlit interface.

    The imported names remain unchanged apart from the requested DDAC ->
    DADMAC display normalization.  Roles always come from the user-confirmed
    IS/SS selections, while targets are sorted by type and chain length.
    """
    roles = resolve_roles(compounds, is_compounds, ss_compounds)
    rows = []
    for role, items in (
        ('Target', sort_compounds(roles['target_compounds'], overrides)),
        ('IS', sort_compounds(roles['is_compounds'], overrides)),
        ('SS', sort_compounds(roles['ss_compounds'], overrides)),
    ):
        for compound in items:
            metadata = compound_metadata_for(compound, overrides)
            rows.append({
                '名称': metadata['name'],
                '类型': metadata['type'],
                '链长': metadata['chain_length'] or 'NA',
                '角色': metadata['role'] or role,
            })
    return rows


def sort_compounds(compounds, overrides=None):
    """Sort detected analytes by recognized family, chain length, then name."""
    family_order = {'BAC': 0, 'DADMAC': 1, 'ATMAC': 2}
    def family_sort_key(name):
        analyte_type = compound_metadata_for(name, overrides)['type']
        if analyte_type in family_order:
            return family_order[analyte_type], ''
        # Non-QAC analytes are arranged by chain length first, then their
        # inferred type.  This keeps e.g. C8-PFOS, C10-Other, C12-PFOS in a
        # practical carbon-chain order while still exposing their types.
        return 3, ''
    return sorted(
        list(dict.fromkeys(compounds)),
        key=lambda name: (
            *family_sort_key(name),
            int(re.sub(r'^C', '', compound_metadata_for(name, overrides)['chain_length']) or 10**9)
            if re.fullmatch(r'C?\d+', compound_metadata_for(name, overrides)['chain_length']) else 10**9,
            compound_metadata_for(name, overrides)['type'],
            str(name).upper(),
        ),
    )


def _cell_text(value):
    return str(value or '').strip()


def _normalise_header_text(value):
    """Normalize harmless header punctuation from CSV/XLSX conversion tools."""
    return re.sub(r'[_\-]+', ' ', _cell_text(value).lower()).strip()


def _is_compound_header(value):
    return _normalise_header_text(value) in {
        'name', 'analyte', 'analyte name', 'compound', 'compound name',
        'compound id', '名称', '化合物', '化合物名称',
    }


def _looks_like_header_row(row):
    texts = [_normalise_header_text(value) for value in row]
    has_name = any(_is_compound_header(value) for value in row)
    has_data_hint = any(('blank' in value or 'sample' in value or 'ms' in value or value.startswith('f')) for value in texts)
    return has_name and has_data_hint


def _csv_rows(raw_bytes):
    # UTF-16 is common when a CSV is opened and re-saved from Excel.  Prefer
    # BOM-aware decoding before legacy East-Asian encodings.
    for encoding in ('utf-8-sig', 'utf-16', 'utf-8', 'gb18030', 'big5'):
        try:
            text = raw_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            text = None
    if text is None:
        raise ValueError('CSV encoding is not supported; please save as UTF-8, UTF-16, or GB18030.')
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t')
    except csv.Error:
        dialect = csv.excel
    return [row for row in csv.reader(io.StringIO(text), dialect)]


def _normalise_rows(rows):
    rows = [[value.strip() if isinstance(value, str) else value for value in row] for row in rows]
    header_index = next((i for i, row in enumerate(rows) if _looks_like_header_row(row)), None)
    if header_index is None:
        header_index = 0
    header = rows[header_index]
    body = rows[header_index + 1:]
    name_col = next((i for i, value in enumerate(header) if _is_compound_header(value)), 0)
    data_start = name_col + 1
    if data_start < len(header) and any(token in _cell_text(header[data_start]).lower()
                                        for token in ('ion', 'transition', '离子', '母离子', '子离子')):
        data_start += 1
    return header, body, name_col, data_start


def _read_csv_source(raw_bytes):
    rows = _csv_rows(raw_bytes)
    header, body, name_col, data_start = _normalise_rows(rows)
    return _parse_tabular_rows(header, body, name_col, data_start)


def _parse_tabular_rows(header, body, compound_col, data_start_col):
    max_cols = max([len(header)] + [len(row) for row in body] or [0])
    header = list(header) + [''] * (max_cols - len(header))
    blanks, mss, samps = [], [], []
    for index in range(data_start_col, max_cols):
        hdr = _cell_text(header[index])
        upper = hdr.upper()
        letter = get_column_letter(index + 1)
        if 'BLANK' in upper:
            blanks.append((index + 1, letter, hdr))
        elif re.search(r'\bMS(?:\d+)?\b|MATRIX\s*SPIKE', upper):
            mss.append((index + 1, letter, hdr))
        elif re.search(r'\d+\s*PPB', upper):
            continue
        elif hdr:
            samps.append((index + 1, letter, hdr))

    data = {}
    compounds = []
    for row in body:
        row = list(row) + [None] * (max_cols - len(row))
        name = normalize_analyte_name(_cell_text(row[compound_col]))
        if not name:
            continue
        compounds.append(name)
        data[name] = {get_column_letter(i + 1): row[i] for i in range(data_start_col, max_cols)}
    target, is_c, ss_c = classify_compounds(compounds)
    return data, blanks, mss, samps, target, is_c, ss_c, target + is_c + ss_c


def resolve_ss_spike(name, cfg, ms_column=None):
    """Resolve a surrogate's own spike concentration by exact analyte name."""
    # The compound-by-MS table is the most specific source when supplied.
    if ms_column is not None:
        configured_matrix = cfg.get('matrix_spike_concentrations') or {}
        _, letter, header = ms_column
        per_ms = configured_matrix.get(name)
        if isinstance(per_ms, dict):
            for key in (header, letter):
                value = safe_float(per_ms.get(key))
                if value is not None and value > 0:
                    return value
    configured = cfg.get('ss_spike_concentrations') or {}
    if name in configured:
        value = safe_float(configured[name])
        if value is not None and value > 0:
            return value
    # Backward-compatible defaults for the current Demo only.
    if 'd7' in name.lower():
        return safe_float(cfg.get('ss_spike_d7_ppb', 4))
    if 'd9' in name.lower():
        return safe_float(cfg.get('ss_spike_d9_ppb', 4))
    return safe_float(cfg.get('ss_spike_conc_ppb'))


def resolve_matrix_spike_concentration(compound, ms_column, cfg):
    """Return one compound's concentration in one matrix-spike column.

    The preferred configuration is ``{compound: {MS1: value, MS2: value}}``.
    A flat header map remains supported for old saved configurations.
    """
    _, letter, header = ms_column
    configured = cfg.get('matrix_spike_concentrations') or {}
    per_ms = configured.get(compound)
    if isinstance(per_ms, dict):
        for key in (header, letter):
            value = safe_float(per_ms.get(key))
            if value is not None and value > 0:
                return value
    for key in (header, letter):
        value = safe_float(configured.get(key))
        if value is not None and value > 0:
            return value
    value = safe_float(cfg.get('spike_conc_ppb', 10))
    if value is None or value <= 0:
        raise ValueError(f'{header}: matrix-spike concentration must be positive.')
    return value


def parse_custom_ss_entries(text):
    """Parse one user-defined surrogate per line: name, spike concentration.

    Example: ``d7-C12-BAC, 4``.  The caller validates that each supplied name
    actually exists in the imported MassHunter table before it is used.
    """
    entries = {}
    errors = []
    for line_number, raw_line in enumerate(str(text or '').splitlines(), 1):
        line = raw_line.strip()
        if not line:
            continue
        parts = [part.strip() for part in re.split(r'[,\t;]', line, maxsplit=1)]
        if len(parts) != 2 or not parts[0]:
            errors.append(f'Line {line_number}: enter "name, concentration".')
            continue
        name = normalize_analyte_name(parts[0])
        concentration = safe_float(parts[1])
        if concentration is None or concentration <= 0:
            errors.append(f'Line {line_number}: spike concentration must be positive.')
            continue
        if name in entries:
            errors.append(f'Line {line_number}: duplicate SS name "{name}".')
            continue
        entries[name] = concentration
    return entries, errors


def parse_compound_name_entries(text):
    """Parse compound names separated by lines, commas, semicolons, or tabs.

    Both English and Chinese comma/semicolon characters are accepted. Names
    are normalized, de-duplicated, and returned in the user's original order.
    """
    names = []
    seen = set()
    for raw_name in re.split(r'[,，;；\t\r\n]+', str(text or '')):
        name = normalize_analyte_name(raw_name.strip())
        if not name or name in seen:
            continue
        seen.add(name)
        names.append(name)
    return names


def missing_matrix_spike_entries(compounds, ms_headers, concentrations):
    """Return compound/MS cells without a positive finite spike value."""
    missing = []
    configured = concentrations or {}
    for compound in compounds or []:
        compound_values = configured.get(compound) or {}
        for header in ms_headers or []:
            value = safe_float(compound_values.get(header))
            if value is None or not math.isfinite(value) or value <= 0:
                missing.append((compound, header))
    return missing


def parse_ss_matrix_spike_entries(text, ms_headers):
    """Parse one SS per line: name followed by one concentration per MS."""
    entries = {}
    errors = []
    headers = list(ms_headers or [])
    for line_number, raw_line in enumerate(str(text or '').splitlines(), 1):
        line = raw_line.strip()
        if not line:
            continue
        line = re.sub(r'[,，;；\t]+\s*$', '', line)
        parts = [part.strip() for part in re.split(r'[,，;；\t]+', line)]
        if not parts[0]:
            errors.append(f'第{line_number}行：缺少替代物名称。')
            continue
        name = normalize_analyte_name(parts[0])
        raw_values = parts[1:]
        if len(raw_values) != len(headers):
            errors.append(
                f'第{line_number}行（{name}）：检测到{len(headers)}个MS，需要{len(headers)}个浓度，实际填写{len(raw_values)}个。'
            )
            continue
        values = [safe_float(value) for value in raw_values]
        if any(value is None or not math.isfinite(value) or value <= 0 for value in values):
            errors.append(f'第{line_number}行（{name}）：每个基质加标浓度必须为大于0的数字。')
            continue
        if name in entries:
            errors.append(f'第{line_number}行：替代物名称“{name}”重复。')
            continue
        entries[name] = dict(zip(headers, values))
    return entries, errors


def one_sided_t99(sample_count):
    """Return t(0.99, n-1) for the available number of replicates."""
    n = int(sample_count or 0)
    if n < 2:
        raise ValueError('At least two replicate results are required for an MDL standard deviation.')
    df = n - 1
    if df in T99_ONE_SIDED:
        return T99_ONE_SIDED[df]
    # Accurate enough for the numeric preview when n is above the table range.
    return 2.32635 + 2.32635 ** 3 / (4 * df)


def calculate_nonzero_blank_mdl(spike_values, blank_values):
    """Calculate mean(blank) + t(0.99,n-1) * SD(blank) in bottle units."""
    blanks = [safe_float(value) for value in blank_values if safe_float(value) is not None]
    if len(blanks) < 2:
        raise ValueError('MDL requires at least two valid blank replicates.')
    return statistics.mean(blanks) + one_sided_t99(len(blanks)) * statistics.stdev(blanks)


def _compact_number(value):
    value = safe_float(value)
    return '' if value is None else f'{value:g}'


def build_blank_mdl_evidence(compound, blank_values, cfg):
    """Return the auditable per-compound evidence used for the vial MDL."""
    raw_values = list(blank_values or [])
    values = [safe_float(value) for value in raw_values if safe_float(value) is not None]
    valid_count = len(values)
    nonzero_count = sum(value != 0 for value in values)
    evidence = {
        'compound': compound,
        'blank_values': values,
        'valid_count': valid_count,
        'nonzero_count': nonzero_count,
        'status': None,
        'ready': False,
        'mean': None,
        'sd': None,
        'degrees_of_freedom': None,
        't_value': None,
        'calibration_concentration': None,
        'signal_to_noise': None,
        'formula': None,
        'mdl': None,
        'reason': None,
    }

    all_cells_zero_or_empty = (
        bool(raw_values)
        and all(is_empty_cell(value) or safe_float(value) == 0 for value in raw_values)
        and nonzero_count == 0
    )
    if all_cells_zero_or_empty:
        evidence['status'] = 'blank_zero'
        override = (cfg.get('mdl_overrides') or {}).get(compound) or {}
        concentration = safe_float(override.get('calibration_concentration'))
        signal_to_noise = safe_float(override.get('signal_to_noise'))
        evidence['calibration_concentration'] = concentration
        evidence['signal_to_noise'] = signal_to_noise
        if concentration is None or concentration <= 0 or signal_to_noise is None or signal_to_noise <= 0:
            evidence['reason'] = 'Zero-or-empty blanks require a positive calibration concentration and S/N.'
            return evidence
        evidence['ready'] = True
        evidence['formula'] = f'3 × {_compact_number(concentration)} ÷ {_compact_number(signal_to_noise)}'
        evidence['mdl'] = 3 * concentration / signal_to_noise
        return evidence


    if valid_count == 0:
        evidence.update(status='missing', reason='No valid blank results were found.')
        return evidence

    if valid_count < 2:
        evidence.update(
            status='insufficient',
            reason='At least two valid blank results are required to calculate a standard deviation.',
        )
        return evidence

    if nonzero_count == 0:
        evidence.update(
            status='incomplete',
            reason='Blank entries contain unsupported text; use genuinely empty cells or numeric zero for the S/N path.',
        )
        return evidence

    mean_value = statistics.mean(values)
    sd_value = statistics.stdev(values)
    t_value = one_sided_t99(valid_count)
    evidence.update(
        status='blank_nonzero',
        ready=True,
        mean=mean_value,
        sd=sd_value,
        degrees_of_freedom=valid_count - 1,
        t_value=t_value,
        formula=(
            f'{_compact_number(mean_value)} + {_compact_number(t_value)} × '
            f'{_compact_number(sd_value)}'
        ),
        mdl=mean_value + t_value * sd_value,
    )
    return evidence


def _excel_array(values):
    return ','.join(f'{safe_float(value):g}' for value in values if safe_float(value) is not None)


def mdl_formula(name, blank_range, cfg):
    """Return the auditable Excel MDL formula for one analyte in bottle units."""
    override = (cfg.get('mdl_overrides') or {}).get(name) or {}
    if override.get('blank_zero'):
        concentration = safe_float(override.get('calibration_concentration'))
        signal_to_noise = safe_float(override.get('signal_to_noise'))
        if concentration is None or concentration <= 0:
            raise ValueError(f'{name}: calibration concentration must be positive for S/N MDL.')
        if signal_to_noise is None or signal_to_noise <= 0:
            raise ValueError(f'{name}: signal-to-noise must be positive for S/N MDL.')
        return f'=3*{concentration}/{signal_to_noise}'
    # TINV(0.02, df) is the legacy two-tailed equivalent of T.INV(0.99, df).
    return f'=AVERAGE({blank_range})+TINV(0.02,COUNT({blank_range})-1)*STDEV({blank_range})'


def mdl_report_formula(name, bottle_ref, cfg):
    """Return MDL in the same sample/report unit as final concentrations."""
    override = (cfg.get('mdl_overrides') or {}).get(name) or {}
    conversion_factor = safe_float(cfg.get('conversion_factor', 1))
    if conversion_factor is None:
        conversion_factor = 1.0
    if override.get('blank_zero'):
        concentration = safe_float(override.get('calibration_concentration'))
        signal_to_noise = safe_float(override.get('signal_to_noise'))
        if concentration is None or concentration <= 0 or signal_to_noise is None or signal_to_noise <= 0:
            raise ValueError(f'{name}: calibration concentration and S/N must be positive.')
        return f'=3*{concentration}/{signal_to_noise}*{conversion_factor:g}'
    return f'={bottle_ref}*{conversion_factor:g}'


def mql_formula(name, blank_range, cfg):
    """Return MQL in bottle units using the selected blank path."""
    override = (cfg.get('mdl_overrides') or {}).get(name) or {}
    if override.get('blank_zero'):
        concentration = safe_float(override.get('calibration_concentration'))
        signal_to_noise = safe_float(override.get('signal_to_noise'))
        if concentration is None or concentration <= 0 or signal_to_noise is None or signal_to_noise <= 0:
            raise ValueError(f'{name}: calibration concentration and S/N must be positive.')
        return f'=10*{concentration}/{signal_to_noise}'
    return f'=AVERAGE({blank_range})+10*STDEV({blank_range})'


def mql_report_formula(name, bottle_ref, blank_range, cfg):
    conversion_factor = safe_float(cfg.get('conversion_factor', 1))
    if conversion_factor is None:
        conversion_factor = 1.0
    return f'={mql_formula(name, blank_range, cfg)[1:]}*{conversion_factor:g}'


def significant_digits_formula(ref, digits=3):
    """Return an Excel formula that displays a value to significant digits."""
    return f'IFERROR(ROUND({ref},{digits}-1-INT(LOG10(ABS({ref})))),0)'


def _round_significant(value, digits=3):
    """Round a numeric value to significant figures without changing its type."""
    value = safe_float(value)
    if value is None or value == 0:
        return 0.0 if value == 0 else None
    places = digits - 1 - int(math.floor(math.log10(abs(value))))
    return round(value, places)


def _format_significant(value, digits=3):
    """Format a number compactly for the human-readable median range."""
    rounded = _round_significant(value, digits)
    if rounded is None:
        return 'NA'
    return f'{rounded:g}'


def _percentile_inc(values, probability):
    """Return the same linear percentile used by Excel PERCENTILE.INC."""
    ordered = sorted(values)
    if not ordered:
        return None
    if len(ordered) == 1:
        return ordered[0]
    position = (len(ordered) - 1) * probability
    lower = int(math.floor(position))
    upper = int(math.ceil(position))
    if lower == upper:
        return ordered[lower]
    fraction = position - lower
    return ordered[lower] + (ordered[upper] - ordered[lower]) * fraction


def compute_analysis_results(raw_data, blank_cols, sample_cols, cfg):
    """Compute the shared numeric source for web previews and XLSX results."""
    cfg = dict(cfg or {})
    cfg['_raw_data'] = raw_data
    conversion_factor = safe_float(cfg.get('conversion_factor', 1))
    if conversion_factor is None:
        conversion_factor = 1.0
    results = {}
    for compound in cfg.get('target_compounds') or []:
        raw_blanks = [raw_data.get(compound, {}).get(column_letter)
                      for _, column_letter, _ in blank_cols]
        evidence = build_blank_mdl_evidence(compound, raw_blanks, cfg)
        if not evidence['ready']:
            raise ValueError(f'{compound}: {evidence["reason"]}')
        blank_average = 0.0 if evidence['status'] == 'blank_zero' else evidence['mean']
        bottle_mdl = evidence['mdl']
        bottle_mql = _preview_mql(compound, blank_cols, cfg)
        report_mdl = bottle_mdl * conversion_factor
        report_mql = bottle_mql * conversion_factor
        half_report_mdl = report_mdl / 2
        sample_values = {}
        detections = {}
        final_values = []
        valid_samples = 0
        numeric_final_samples = 0
        true_detections = 0
        for _, column_letter, header in sample_cols:
            value = safe_float(raw_data.get(compound, {}).get(column_letter))
            if value is None:
                sample_values[header] = None
                detections[header] = None
                continue
            valid_samples += 1
            detected = value >= bottle_mdl
            detections[header] = 1 if detected else 0
            if detected:
                true_detections += 1
                final_value = (value - blank_average) * conversion_factor
            else:
                final_value = half_report_mdl
            final_value = round6(final_value)
            sample_values[header] = final_value
            final_values.append(final_value)
            numeric_final_samples += 1

        df_fraction = numeric_final_samples / len(sample_cols) if sample_cols else 0.0
        statistics_eligible = df_fraction > 0.5
        if final_values and statistics_eligible:
            statistics_values = {
                'mean': statistics.mean(final_values),
                'geomean': statistics.geometric_mean(final_values) if all(value > 0 for value in final_values) else None,
                'median': statistics.median(final_values),
                'min': min(final_values),
                'max': max(final_values),
                'p05': _percentile_inc(final_values, 0.05),
                'p25': _percentile_inc(final_values, 0.25),
                'p75': _percentile_inc(final_values, 0.75),
                'p95': _percentile_inc(final_values, 0.95),
            }
        else:
            statistics_values = {key: None for key in ('mean', 'geomean', 'median', 'min', 'max', 'p05', 'p25', 'p75', 'p95')}
        results[compound] = {
            'blank_average': blank_average,
            'bottle_mdl': bottle_mdl,
            'bottle_mql': bottle_mql,
            'report_mdl': report_mdl,
            'report_mql': report_mql,
            'half_report_mdl': half_report_mdl,
            'sample_values': sample_values,
            'detections': detections,
            'valid_samples': valid_samples,
            'true_detections': true_detections,
            # Old template rule: numeric final-result cells divided by all
            # nominal sample columns. Missing source cells remain in the denominator.
            'df_fraction': df_fraction,
            'statistics_eligible': statistics_eligible,
            'final_values': final_values,
            **statistics_values,
        }
    return results


def cache_formula_value(cfg, worksheet, cell, value):
    cfg.setdefault('_formula_cache', {})[(worksheet.title, cell.coordinate)] = value


def inject_formula_cached_values(xlsx_bytes, cached_values):
    """Add cached formula results without removing auditable formulas."""
    if not cached_values:
        return xlsx_bytes
    spreadsheet_ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    document_rel_ns = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    package_rel_ns = 'http://schemas.openxmlformats.org/package/2006/relationships'
    ET.register_namespace('', spreadsheet_ns)
    ET.register_namespace('r', document_rel_ns)
    source = io.BytesIO(xlsx_bytes)
    destination = io.BytesIO()
    with zipfile.ZipFile(source, 'r') as zin:
        workbook_root = ET.fromstring(zin.read('xl/workbook.xml'))
        relationships = ET.fromstring(zin.read('xl/_rels/workbook.xml.rels'))
        targets = {rel.attrib['Id']: rel.attrib['Target']
                   for rel in relationships.findall(f'{{{package_rel_ns}}}Relationship')}
        sheet_paths = {}
        for sheet in workbook_root.find(f'{{{spreadsheet_ns}}}sheets'):
            rel_id = sheet.attrib[f'{{{document_rel_ns}}}id']
            target = targets[rel_id].lstrip('/')
            sheet_paths[sheet.attrib['name']] = target if target.startswith('xl/') else f'xl/{target}'
        by_path = {}
        for (sheet_name, coordinate), value in cached_values.items():
            path = sheet_paths.get(sheet_name)
            if path:
                by_path.setdefault(path, {})[coordinate] = value
        with zipfile.ZipFile(destination, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                updates = by_path.get(item.filename)
                if updates:
                    root = ET.fromstring(data)
                    for cell in root.findall(f'.//{{{spreadsheet_ns}}}c'):
                        coordinate = cell.attrib.get('r')
                        if coordinate not in updates:
                            continue
                        value = updates[coordinate]
                        value_node = cell.find(f'{{{spreadsheet_ns}}}v')
                        if value_node is None:
                            value_node = ET.SubElement(cell, f'{{{spreadsheet_ns}}}v')
                        if value is None:
                            value_node.text = None
                        elif isinstance(value, str):
                            cell.set('t', 'str')
                            value_node.text = value
                        else:
                            cell.attrib.pop('t', None)
                            value_node.text = f'{float(value):.15g}'
                    data = ET.tostring(root, encoding='utf-8', xml_declaration=True)
                zout.writestr(item, data)
    return destination.getvalue()


def _numeric_values(raw_data, compound, columns):
    return [safe_float(raw_data.get(compound, {}).get(column_letter))
            for _, column_letter, _ in columns
            if safe_float(raw_data.get(compound, {}).get(column_letter)) is not None]


def detect_blank_zero_compounds(raw_data, blank_cols):
    """Detect analytes whose blank series contains only numeric zero or empty cells."""
    detected = []
    for compound in raw_data:
        values = [raw_data.get(compound, {}).get(column_letter)
                  for _, column_letter, _ in blank_cols]
        if values and all(is_empty_cell(value) or safe_float(value) == 0 for value in values):
            detected.append(compound)
    return detected


def validate_blank_zero_mdl(compound, blank_values, cfg):
    """Validate that a zero-or-empty blank analyte has the required manual S/N rule."""
    values = list(blank_values or [])
    if not values or not all(is_empty_cell(value) or safe_float(value) == 0 for value in values):
        return None
    override = (cfg.get('mdl_overrides') or {}).get(compound) or {}
    concentration = safe_float(override.get('calibration_concentration'))
    signal_to_noise = safe_float(override.get('signal_to_noise'))
    if not override.get('blank_zero') or concentration is None or concentration <= 0 or signal_to_noise is None or signal_to_noise <= 0:
        raise ValueError(f'{compound}: zero-or-empty blanks require positive calibration concentration and S/N.')
    return None


def validate_blank_zero_configuration(raw_data, blank_cols, target_compounds, cfg):
    """Reject processing when any target lacks enough MDL calculation evidence."""
    for compound in target_compounds:
        blank_values = [raw_data.get(compound, {}).get(column_letter)
                        for _, column_letter, _ in blank_cols]
        evidence = build_blank_mdl_evidence(compound, blank_values, cfg)
        if not evidence['ready']:
            raise ValueError(f'{compound}: {evidence["reason"]}')


def _preview_mdl(compound, blank_cols, cfg):
    blank_values = [cfg.get('_raw_data', {}).get(compound, {}).get(column_letter)
                    for _, column_letter, _ in blank_cols]
    evidence = build_blank_mdl_evidence(compound, blank_values, cfg)
    return evidence['mdl'] if evidence['ready'] else None


def _preview_mql(compound, blank_cols, cfg):
    override = (cfg.get('mdl_overrides') or {}).get(compound) or {}
    if override.get('blank_zero'):
        concentration = safe_float(override.get('calibration_concentration'))
        signal_to_noise = safe_float(override.get('signal_to_noise'))
        if concentration is None or concentration <= 0 or signal_to_noise is None or signal_to_noise <= 0:
            raise ValueError(f'{compound}: calibration concentration and S/N must be positive.')
        return 10 * concentration / signal_to_noise
    blanks = _numeric_values(cfg.get('_raw_data', {}), compound, blank_cols)
    if len(blanks) < 2:
        return None
    return statistics.mean(blanks) + 10 * statistics.stdev(blanks)


def compute_preview_summary(raw_data, blank_cols, sample_cols, cfg):
    """Calculate numeric summary values for Streamlit without relying on Excel recalculation."""
    analysis_results = compute_analysis_results(raw_data, blank_cols, sample_cols, cfg)
    rows = []
    compounds = cfg.get('target_compounds') or []
    for compound in compounds:
        analysis = analysis_results[compound]
        if analysis['statistics_eligible']:
            median_iqr = (
                f'{_format_significant(analysis["median"])} '
                f'({_format_significant(analysis["p25"])}-{_format_significant(analysis["p75"])})'
            )
        else:
            median_iqr = 'NC'

        rows.append({
            '名称': compound,
            '链长': extract_chain_length(compound) or 'NA',
            'DF (%)': _round_significant(analysis['df_fraction'] * 100),
            'Median (Q1-Q3)': median_iqr,
            'MDL': _round_significant(analysis['report_mdl']),
            'MQL': _round_significant(analysis['report_mql']),
        })
    return rows


def compute_preview_final_table(raw_data, blank_cols, sample_cols, cfg):
    """Return numeric final-concentration rows for an online sample preview."""
    analysis_results = compute_analysis_results(raw_data, blank_cols, sample_cols, cfg)
    rows = []
    for compound in cfg.get('target_compounds') or []:
        analysis = analysis_results[compound]
        result = {'名称': compound}
        for _, _, header in sample_cols:
            result[header] = analysis['sample_values'].get(header)
        rows.append(result)
    return rows


# ============================================================
# 样式
# ============================================================
def make_styles():
    base = Font(name='宋体', size=11, color=FONT_COLOR)
    bold = Font(name='宋体', size=11, color=FONT_COLOR, bold=True)
    red_bold = Font(name='宋体', size=11, color=RED_FONT, bold=True)
    blue = Font(name='宋体', size=11, color=BLUE_FONT)
    gold = Font(name='宋体', size=11, color=GOLD_FONT)

    hdr_fill  = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type='solid')
    blue_fill = PatternFill(start_color=BLUE_BG, end_color=BLUE_BG, fill_type='solid')
    gold_fill = PatternFill(start_color=GOLD_BG, end_color=GOLD_BG, fill_type='solid')
    yell_fill = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type='solid')
    no_fill   = PatternFill(fill_type=None)

    ca = Alignment(horizontal='center', vertical='center')
    la = Alignment(horizontal='left', vertical='center')

    return {
        'hdr':  {'font': base, 'fill': hdr_fill,  'alignment': ca},
        'hdrL': {'font': base, 'fill': hdr_fill,  'alignment': la},
        'cmpd': {'font': base, 'fill': no_fill,   'alignment': la},
        'data': {'font': base, 'fill': no_fill,   'alignment': ca},
        'rec':  {'font': red_bold, 'fill': no_fill, 'alignment': ca},
        'stat': {'font': blue, 'fill': blue_fill, 'alignment': ca},
        'statB':{'font': blue, 'fill': blue_fill, 'alignment': ca},  # stat header bold
        'gold': {'font': gold, 'fill': gold_fill, 'alignment': ca},
        'yell': {'font': base, 'fill': yell_fill, 'alignment': ca},
        'yellL':{'font': bold, 'fill': yell_fill, 'alignment': la},
        'bold': {'font': bold, 'fill': no_fill,   'alignment': ca},
    }

def sty(cell, s):  # apply style dict
    for k in ('font','fill','alignment'):
        if s.get(k): setattr(cell, k, s[k])


# ============================================================
# 原始数据读取
# ============================================================
def read_raw(filepath_or_bytes):
    if isinstance(filepath_or_bytes, (bytes, bytearray)):
        raw_bytes = bytes(filepath_or_bytes)
        if not raw_bytes.startswith(b'PK'):
            if raw_bytes.startswith(b'\xd0\xcf\x11\xe0'):
                frame = pd.read_excel(io.BytesIO(raw_bytes), header=None, engine='xlrd')
                rows = frame.where(pd.notna(frame), None).values.tolist()
                header, body, compound_col, data_start = _normalise_rows(rows)
                return _parse_tabular_rows(header, body, compound_col, data_start)
            return _read_csv_source(raw_bytes)
    if isinstance(filepath_or_bytes, bytes):
        wb = openpyxl.load_workbook(io.BytesIO(filepath_or_bytes), data_only=True)
    else:
        if os.fspath(filepath_or_bytes).lower().endswith('.xls'):
            frame = pd.read_excel(filepath_or_bytes, header=None, engine='xlrd')
            rows = frame.where(pd.notna(frame), None).values.tolist()
            header, body, compound_col, data_start = _normalise_rows(rows)
            return _parse_tabular_rows(header, body, compound_col, data_start)
        wb = openpyxl.load_workbook(filepath_or_bytes, data_only=True)
    ws = next((candidate for candidate in wb.worksheets
               if any(_looks_like_header_row(list(row))
                      for row in candidate.iter_rows(min_row=1, max_row=min(candidate.max_row, 20), values_only=True))), None)
    if ws is None and 'Sheet1' in wb.sheetnames:
        ws = wb['Sheet1']
    if ws is None:
        ws = next((candidate for candidate in wb.worksheets
                   if candidate.max_row > 0 and candidate.max_column > 0), None)
    if ws is None:
        wb.close()
        raise ValueError('No non-empty worksheet was found.')

    preview_rows = [list(row) for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True)]
    header_index = next((i for i, row in enumerate(preview_rows) if _looks_like_header_row(row)), None)
    if header_index is not None:
        # The header is detected from a small preview, but the body must be
        # read through the worksheet's actual last row.
        rows = [list(row) for row in ws.iter_rows(min_row=1, values_only=True)]
        header, body, compound_col, data_start = _normalise_rows(rows)
        result = _parse_tabular_rows(header, body, compound_col, data_start)
        wb.close()
        return result

    compound_col = 2; data_start_col = 4
    for col in range(1, ws.max_column + 1):
        v = str(ws.cell(row=2, column=col).value or '').strip()
        if v == '名称':
            compound_col = col; data_start_col = col + 1
            next_v = str(ws.cell(row=2, column=col+1).value or '').strip()
            if '离子' in next_v or next_v == '':
                data_start_col = col + 2 if col + 2 <= ws.max_column else col + 1
            break

    blanks, mss, samps = [], [], []
    for col in range(data_start_col, ws.max_column + 1):
        hdr = str(ws.cell(row=1, column=col).value or '')
        cl = get_column_letter(col)
        if 'BLANK' in hdr.upper():
            blanks.append((col, cl, hdr))
        elif 'MS' in hdr.upper():
            mss.append((col, cl, hdr))
        elif re.search(r'\dPPB', hdr.upper()):
            continue
        else:
            samps.append((col, cl, hdr))

    data = {}
    for row in range(3, ws.max_row + 1):
        nm = normalize_analyte_name(str(ws.cell(row=row, column=compound_col).value or '').strip())
        if not nm: continue
        data[nm] = {}
        for col in range(data_start_col, ws.max_column + 1):
            data[nm][get_column_letter(col)] = ws.cell(row=row, column=col).value
    # 读取化合物名称用于分类
    compounds_raw = []
    for row in range(3, ws.max_row + 1):
        nm = normalize_analyte_name(str(ws.cell(row=row, column=compound_col).value or '').strip())
        if nm: compounds_raw.append(nm)
    wb.close()

    target, is_c, ss_c = classify_compounds(compounds_raw)
    return data, blanks, mss, samps, target, is_c, ss_c, target + is_c + ss_c


def validate_input_layout(blanks, mss, samps, target_compounds, is_compounds, ss_compounds):
    """Return a human-readable compatibility check for an imported MassHunter file."""
    errors = []
    warnings = []
    if not blanks:
        errors.append('未识别到 BLANK 列；请检查列名是否包含 BLANK。')
    if not samps:
        errors.append('未识别到 sample 列；样品列应使用 F1、F2 或 Sample-1 等名称。')
    if not target_compounds:
        errors.append('未识别到目标化合物行；请检查化合物名称列和数据行。')
    if not mss:
        warnings.append('未识别到 MS/基质加标列；基质加标回收率将无法计算。')
    if len(blanks) < 2:
        warnings.append('BLANK 列少于 2 个；标准差和 MDL 的稳定性需要确认。')
    if not is_compounds:
        warnings.append('未识别到 IS 内标；请在界面中确认是否使用内标校正。')
    if not ss_compounds:
        warnings.append('未识别到 SS 替代物；SS 回收率不会自动生成。')
    all_headers = [item[2].strip().upper() for item in [*blanks, *mss, *samps] if item[2]]
    duplicate_headers = sorted({header for header in all_headers if all_headers.count(header) > 1})
    if duplicate_headers:
        warnings.append('发现重复数据列名：' + ', '.join(duplicate_headers))
    return {
        'ready': not errors,
        'errors': errors,
        'warnings': warnings,
        'summary': f'{len(blanks)} BLANK + {len(mss)} MS + {len(samps)} sample + '
                   f'{len(is_compounds)} IS + {len(ss_compounds)} SS + {len(target_compounds)} 个目标物',
    }


# ============================================================
# 化合物自动分类
# ============================================================
def classify_compounds(compounds):
    bac, ddac, atmac, metabolites, is_list, ss_list, others = [], [], [], [], [], [], []
    for c in list(dict.fromkeys(str(value).strip() for value in compounds if str(value).strip())):
        cn = normalize_analyte_name(c)
        if not cn: continue
        lower = cn.lower()
        if (any(kw in cn for kw in ['[C13]', '-d6', '-d3', 'OH-d3'])
                or re.search(r'\binternal[ _-]?standard\b|\bis\b', lower)):
            is_list.append(cn)
        elif re.match(r'd\d+-', cn) or re.search(r'\bsurrogate\b|\bss\b', lower):
            ss_list.append(cn)
        elif 'DADMAC' in cn.upper(): ddac.append(cn)
        elif 'ATMAC' in cn.upper(): atmac.append(cn)
        elif 'BAC' in cn.upper(): bac.append(cn)
        else: others.append(cn)
    def sk(n):
        nums = re.findall(r'C(\d+)', n)
        return int(nums[0]) if nums else 0
    bac.sort(key=sk); ddac.sort(key=sk); atmac.sort(key=sk)
    metabolites.sort(key=sk); is_list.sort(key=sk); ss_list.sort(key=sk)
    pure_bac = []; meta = []
    for c in bac:
        if '+O' in c or '+2O' in c: meta.append(c)
        else: pure_bac.append(c)
    metabolites = meta + metabolites
    target = sort_compounds(pure_bac + ddac + atmac + metabolites + others)
    return target, is_list, ss_list


# ============================================================
# Sheet 1: 基质加标浓度
# ============================================================
def build_sheet1(wb, raw_data, ms_cols, S, cfg):
    text = workbook_text(cfg)
    ws = wb.create_sheet(text['matrix'])
    n_ms = len(ms_cols)
    # 列布局: A | B~(n_ms) MS data | 空 | Recoveries标签 | n_ms个回收率% | 空 | avg | SD | SE
    # SS 的理论加标浓度仅作为处理参数，用于其独立回收率计算；不增加输出列。
    ms_start = 2
    mid1 = ms_start + n_ms           # 空列
    rec_lbl = mid1 + 1               # "Recoveries"
    rec_start = rec_lbl + 1          # 回收率% 起始
    mid2 = rec_start + n_ms          # 空列
    stat_col = mid2 + 1              # avg
    sd_col = stat_col + 1            # SD
    se_col = stat_col + 2            # SE
    last_col = se_col

    # Row 1
    ws.cell(row=1, column=1, value=text['compound']); sty(ws.cell(row=1,column=1), S['hdrL'])
    for i in range(n_ms):
        ws.cell(row=1, column=ms_start+i, value=f'matrix spike_{i+1}'); sty(ws.cell(row=1,column=ms_start+i), S['hdr'])
    ws.cell(row=1, column=mid1, value=None)
    ws.cell(row=1, column=rec_lbl, value='Recoveries'); sty(ws.cell(row=1,column=rec_lbl), S['hdr'])
    for i in range(n_ms):
        ws.cell(row=1, column=rec_start+i, value=f'matrix spike_{i+1}'); sty(ws.cell(row=1,column=rec_start+i), S['hdr'])
    ws.cell(row=1, column=mid2, value=None)
    ws.cell(row=1, column=stat_col, value='average'); sty(ws.cell(row=1,column=stat_col), S['hdr'])
    ws.cell(row=1, column=sd_col, value='SD'); sty(ws.cell(row=1,column=sd_col), S['hdr'])
    ws.cell(row=1, column=se_col, value='SE'); sty(ws.cell(row=1,column=se_col), S['hdr'])

    # Row 2
    ws.cell(row=2, column=1, value=text['group']); sty(ws.cell(row=2,column=1), S['hdr'])
    for i in range(n_ms):
        ws.cell(row=2, column=ms_start+i, value='ppb'); sty(ws.cell(row=2,column=ms_start+i), S['hdr'])
    ws.cell(row=2, column=mid1, value=None)
    ws.cell(row=2, column=rec_lbl, value=None)
    for i in range(n_ms):
        ws.cell(row=2, column=rec_start+i, value='%'); sty(ws.cell(row=2,column=rec_start+i), S['hdr'])
    ws.cell(row=2, column=mid2, value=None)
    ws.cell(row=2, column=stat_col, value='%'); sty(ws.cell(row=2,column=stat_col), S['hdr'])
    ws.cell(row=2, column=sd_col, value='%'); sty(ws.cell(row=2,column=sd_col), S['hdr'])
    ws.cell(row=2, column=se_col, value='%'); sty(ws.cell(row=2,column=se_col), S['hdr'])

    rec_cl = get_column_letter(rec_start)
    rec_cr = get_column_letter(rec_start + n_ms - 1)

    # 数据行（跳过 SS，SS 单独放底部）
    row = 3
    all_compounds = cfg.get('all_compounds', ALL_COMPS)
    selected_is = set(cfg.get('is_compounds', IS_COMPS))
    selected_ss = cfg.get('ss_compounds', SS_COMPS)
    non_ss = [c for c in all_compounds if c not in selected_ss and c not in selected_is]
    for comp in non_ss:
        ws.cell(row=row, column=1, value=comp); sty(ws.cell(row=row,column=1), S['cmpd'])

        # MS 原始数据
        for i, (_, cl, _) in enumerate(ms_cols):
            v = safe_float(raw_data.get(comp, {}).get(cl))
            c = ws.cell(row=row, column=ms_start+i)
            if v is not None:
                c.value = round6(v); c.number_format = '0.000000'
            sty(c, S['data'])

        ws.cell(row=row, column=mid1, value=None)
        ws.cell(row=row, column=rec_lbl, value=None)

        # 回收率 % = (MS值/spike)*100, 四舍五入取整
        for i, ms_col in enumerate(ms_cols):
            _, cl, _ = ms_col
            v = safe_float(raw_data.get(comp, {}).get(cl))
            c = ws.cell(row=row, column=rec_start+i)
            if v is not None:
                c.value = round_int((v / resolve_matrix_spike_concentration(comp, ms_col, cfg)) * 100)
                c.number_format = '0'
            sty(c, S['rec'])

        ws.cell(row=row, column=mid2, value=None)

        # 统计公式 (基于回收率%列)
        recovery_values = []
        for ms_col in ms_cols:
            _, column_letter, _ = ms_col
            measured = safe_float(raw_data.get(comp, {}).get(column_letter))
            if measured is not None:
                recovery_values.append(round_int(
                    measured / resolve_matrix_spike_concentration(comp, ms_col, cfg) * 100
                ))
        rr = f'{rec_cl}{row}:{rec_cr}{row}'
        c_avg = ws.cell(row=row, column=stat_col)
        c_avg.value = f'=ROUND(AVERAGE({rr}),0)'; c_avg.number_format = '0'
        sty(c_avg, S['data'])
        cache_formula_value(cfg, ws, c_avg, round_int(statistics.mean(recovery_values)) if recovery_values else None)

        c_sd = ws.cell(row=row, column=sd_col)
        c_sd.value = f'=ROUND(STDEV({rr}),0)'; c_sd.number_format = '0'
        sty(c_sd, S['data'])
        recovery_sd = round_int(statistics.stdev(recovery_values)) if len(recovery_values) >= 2 else None
        cache_formula_value(cfg, ws, c_sd, recovery_sd)

        c_se = ws.cell(row=row, column=se_col)
        c_se.value = f'=ROUND({get_column_letter(sd_col)}{row}/SQRT(COUNT({rr})),0)'; c_se.number_format = '0'
        sty(c_se, S['data'])
        cache_formula_value(
            cfg, ws, c_se,
            round_int(recovery_sd / math.sqrt(len(recovery_values))) if recovery_sd is not None else None,
        )

        row += 1

    # --- SS 替代物回收率部分（独立于上方主列表） ---
    row += 1
    ws.cell(row=row, column=1, value=text['ss_recoveries'])
    sty(ws.cell(row=row,column=1), S['yellL'])
    for c in range(2, last_col+1):
        sty(ws.cell(row=row,column=c), S['yell'])
    row += 1

    selected_ss = cfg.get('ss_compounds', SS_COMPS)
    for ss in selected_ss:
        ws.cell(row=row, column=1, value=ss); sty(ws.cell(row=row,column=1), S['cmpd'])
        # MS数据列：照抄原始浓度（不除以任何值）
        for i, (_, cl, _) in enumerate(ms_cols):
            v = safe_float(raw_data.get(ss, {}).get(cl))
            c = ws.cell(row=row, column=ms_start+i)
            if v is not None:
                c.value = round6(v); c.number_format = '0.000000'
            sty(c, S['data'])
        ws.cell(row=row, column=mid1, value=None)
        ws.cell(row=row, column=rec_lbl, value=None)
        # 回收率列：SS实测浓度 ÷ SS理论加标浓度 × 100%
        for i, ms_col in enumerate(ms_cols):
            _, cl, _ = ms_col
            v = safe_float(raw_data.get(ss, {}).get(cl))
            c = ws.cell(row=row, column=rec_start+i)
            if v is not None:
                this_ss_spike = resolve_ss_spike(ss, cfg, ms_col)
                if this_ss_spike is None or this_ss_spike <= 0:
                    raise ValueError(f'{ss}: missing positive SS spike concentration for {ms_col[2]}.')
                c.value = round_int((v / this_ss_spike) * 100)
                c.number_format = '0'
            sty(c, S['rec'])
        for c in range(mid2, last_col+1):
            sty(ws.cell(row=row,column=c), S['data'])
        recovery_values = []
        for ms_col in ms_cols:
            _, column_letter, _ = ms_col
            measured = safe_float(raw_data.get(ss, {}).get(column_letter))
            if measured is not None:
                spike = resolve_ss_spike(ss, cfg, ms_col)
                recovery_values.append(round_int(measured / spike * 100))
        rr = f'{rec_cl}{row}:{rec_cr}{row}'
        c_avg = ws.cell(row=row, column=stat_col)
        c_avg.value = f'=ROUND(AVERAGE({rr}),0)'; c_avg.number_format = '0'
        sty(c_avg, S['data'])
        cache_formula_value(cfg, ws, c_avg, round_int(statistics.mean(recovery_values)) if recovery_values else None)
        c_sd = ws.cell(row=row, column=sd_col)
        c_sd.value = f'=ROUND(STDEV({rr}),0)'; c_sd.number_format = '0'
        sty(c_sd, S['data'])
        recovery_sd = round_int(statistics.stdev(recovery_values)) if len(recovery_values) >= 2 else None
        cache_formula_value(cfg, ws, c_sd, recovery_sd)
        c_se = ws.cell(row=row, column=se_col)
        c_se.value = f'=ROUND({get_column_letter(sd_col)}{row}/SQRT(COUNT({rr})),0)'; c_se.number_format = '0'
        sty(c_se, S['data'])
        cache_formula_value(
            cfg, ws, c_se,
            round_int(recovery_sd / math.sqrt(len(recovery_values))) if recovery_sd is not None else None,
        )
        row += 1

    # IS is deliberately separate from recoveries. Its measured concentrations
    # are copied from the original MS columns and never divided into recoveries.
    selected_is = cfg.get('is_compounds', IS_COMPS)
    if selected_is:
        row += 1
        ws.cell(row=row, column=1, value=text['is_additions'])
        sty(ws.cell(row=row, column=1), S['yellL'])
        for c in range(2, last_col + 1):
            sty(ws.cell(row=row, column=c), S['yell'])
        row += 1
        for is_name in selected_is:
            ws.cell(row=row, column=1, value=is_name); sty(ws.cell(row=row, column=1), S['cmpd'])
            for i, ms_col in enumerate(ms_cols):
                c = ws.cell(row=row, column=ms_start + i)
                _, column_letter, _ = ms_col
                value = safe_float(raw_data.get(is_name, {}).get(column_letter))
                if value is not None:
                    c.value = round6(value); c.number_format = '0.000000'
                sty(c, S['data'])
            for c in range(mid1, last_col + 1):
                sty(ws.cell(row=row, column=c), S['data'])
            row += 1

    # 说明
    row += 1
    note = rec_lbl + n_ms
    ws.cell(row=row, column=note, value='此表格计算方法：回收率，用测得浓度除以加标浓度')
    sty(ws.cell(row=row,column=note), S['yell'])
    row += 1
    ws.cell(row=row, column=note, value='Each target or SS recovery uses its own corresponding MS spike amount. IS measured concentrations are copied from the raw MS columns and have no recovery calculation.' if cfg.get('language') == 'en' else '每个目标物或SS回收率均使用其对应MS加标浓度；IS实测浓度来自原始MS列，仅展示，不计算回收率。')
    sty(ws.cell(row=row,column=note), S['yell'])

    ws.row_dimensions[1].height = 19.5
    ws.row_dimensions[2].height = 17.25
    ws.column_dimensions['A'].width = 30.0

    return ws


# ============================================================
# Sheet 2: 空白基质检出限
# ============================================================
def build_sheet2(wb, raw_data, blank_cols, S, cfg):
    text = workbook_text(cfg)
    ws = wb.create_sheet(text['blank'])
    n_b = len(blank_cols)
    cf = cfg['conversion_factor']
    unit = cfg['output_unit']

    # 换算因子存放在 A1
    ws.cell(row=1, column=1, value=cf)
    ws.cell(row=1, column=1).number_format = '0.000000'

    # Row 1 说明
    ws.cell(row=1, column=9, value=text['blank_note'])
    sty(ws.cell(row=1,column=9), S['yell'])

    # Row 2 表头
    ws.cell(row=2, column=1, value=text['compound']); sty(ws.cell(row=2,column=1), S['hdr'])
    for i in range(n_b):
        ws.cell(row=2, column=2+i, value=f'blank_{i+1}'); sty(ws.cell(row=2,column=2+i), S['hdr'])

    blank_end = 2 + n_b         # 空列
    avg_c = blank_end + 1       # I: average
    mdl_c = blank_end + 2       # J: MDL
    half_c = blank_end + 3      # K: 1/2 MDL
    last_c = half_c

    ws.cell(row=2, column=avg_c, value='procedural blank average'); sty(ws.cell(row=2,column=avg_c), S['hdr'])
    ws.cell(row=2, column=mdl_c, value='MDL'); sty(ws.cell(row=2,column=mdl_c), S['hdr'])
    ws.cell(row=2, column=half_c, value='1/2 MDL'); sty(ws.cell(row=2,column=half_c), S['hdr'])

    # Row 3 单位
    ws.cell(row=3, column=1, value=text['group']); sty(ws.cell(row=3,column=1), S['hdr'])
    for i in range(n_b):
        ws.cell(row=3, column=2+i, value='ppb'); sty(ws.cell(row=3,column=2+i), S['hdr'])
    ws.cell(row=3, column=blank_end, value=None)
    ws.cell(row=3, column=avg_c, value='bottle ppb'); sty(ws.cell(row=3,column=avg_c), S['hdr'])
    ws.cell(row=3, column=mdl_c, value='bottle ppb'); sty(ws.cell(row=3,column=mdl_c), S['hdr'])
    ws.cell(row=3, column=half_c, value=unit); sty(ws.cell(row=3,column=half_c), S['hdr'])

    # Row 4 样本标签
    row = 4
    ws.cell(row=row, column=1, value=cfg['sample_type']); sty(ws.cell(row=row,column=1), S['cmpd'])
    for c in range(2, last_c+1): sty(ws.cell(row=row,column=c), S['data'])
    row += 1

    avg_l = get_column_letter(avg_c)
    blank_l = get_column_letter(2)
    blank_r = get_column_letter(2 + n_b - 1)

    first_data_row = row
    target_compounds = cfg.get('target_compounds', TARGET_COMPS)
    for comp in target_compounds:
        analysis = cfg['_analysis_results'][comp]
        ws.cell(row=row, column=1, value=comp); sty(ws.cell(row=row,column=1), S['cmpd'])

        # Blank 值
        for i, (_, cl, _) in enumerate(blank_cols):
            v = safe_float(raw_data.get(comp, {}).get(cl))
            c = ws.cell(row=row, column=2+i)
            if v is not None:
                c.value = round6(v); c.number_format = '0.000000'
            sty(c, S['data'])

        ws.cell(row=row, column=blank_end, value=None)

        # I: Average (公式)
        br = f'{blank_l}{row}:{blank_r}{row}'
        c_avg = ws.cell(row=row, column=avg_c)
        c_avg.value = '=0' if analysis['blank_average'] == 0 and (cfg.get('mdl_overrides') or {}).get(comp, {}).get('blank_zero') else f'=AVERAGE({br})'
        c_avg.number_format = '0.000000'
        sty(c_avg, S['data'])
        cache_formula_value(cfg, ws, c_avg, analysis['blank_average'])

        # J: MDL = 3*STDEVA(blanks) — 公式
        c_mdl = ws.cell(row=row, column=mdl_c)
        c_mdl.value = mdl_formula(comp, br, cfg)
        c_mdl.number_format = '0.000000'
        sty(c_mdl, S['data'])
        cache_formula_value(cfg, ws, c_mdl, analysis['bottle_mdl'])

        # K: 1/2 MDL = ROUND(MDL/2 * $A$1, 6) — 公式，引用A1换算因子
        mdl_l = get_column_letter(mdl_c)
        c_half = ws.cell(row=row, column=half_c)
        c_half.value = f'=ROUND({mdl_l}{row}/2*$A$1,6)'
        c_half.number_format = '0.000000'
        sty(c_half, S['data'])
        cache_formula_value(cfg, ws, c_half, analysis['half_report_mdl'])

        row += 1

    ws.row_dimensions[1].height = 19.5
    ws.row_dimensions[2].height = 17.25
    ws.column_dimensions['A'].width = 30.0

    info = {
        'avg_c': avg_c, 'mdl_c': mdl_c, 'half_c': half_c,
        'avg_l': avg_l, 'mdl_l': get_column_letter(mdl_c), 'half_l': get_column_letter(half_c),
        'first_row': first_data_row,
        'row_map': {comp: first_data_row + index for index, comp in enumerate(target_compounds)},
    }
    return ws, info


# ============================================================
# Sheet 3: 瓶内实测浓度
# ============================================================
def build_sheet3(wb, raw_data, sample_cols, S, cfg):
    text = workbook_text(cfg)
    ws = wb.create_sheet(text['bottle'])
    n_s = len(sample_cols)

    ws.cell(row=1, column=1, value=text['compound']); sty(ws.cell(row=1,column=1), S['hdr'])
    for i, (_, _, hdr) in enumerate(sample_cols):
        ws.cell(row=1, column=2+i, value=hdr.replace('0527 Urine-',''))
        sty(ws.cell(row=1,column=2+i), S['hdr'])

    ws.cell(row=2, column=1, value=text['group']); sty(ws.cell(row=2,column=1), S['hdr'])
    for i in range(n_s):
        ws.cell(row=2, column=2+i, value=cfg['masshunter_unit']); sty(ws.cell(row=2,column=2+i), S['hdr'])

    row = 3
    ws.cell(row=row, column=1, value=cfg['sample_type']); sty(ws.cell(row=row,column=1), S['cmpd'])
    for i in range(n_s): sty(ws.cell(row=row,column=2+i), S['data'])
    row += 1
    first_row = row

    for comp in cfg.get('target_compounds', TARGET_COMPS):
        ws.cell(row=row, column=1, value=comp); sty(ws.cell(row=row,column=1), S['cmpd'])
        for i, (_, cl, _) in enumerate(sample_cols):
            v = safe_float(raw_data.get(comp, {}).get(cl))
            c = ws.cell(row=row, column=2+i)
            if v is not None:
                c.value = round6(v); c.number_format = '0.000000'
            sty(c, S['data'])
        row += 1

    ws.row_dimensions[1].height = 19.5
    ws.row_dimensions[2].height = 17.25
    ws.column_dimensions['A'].width = 30.0
    return ws, first_row


# ============================================================
# Sheet 4: 最终计算浓度 (v2: 删除D/E列, 合并统计表头, A38标注)
# ============================================================
def build_sheet4(wb, raw_data, sample_cols, blank_info, s3_first, S, cfg):
    text = workbook_text(cfg)
    ws = wb.create_sheet(text['final'])
    n_s = len(sample_cols)
    cf = cfg['conversion_factor']
    unit = cfg['output_unit']

    # 列布局 (删除D/E后):
    # A:化合物 B:BLANK avg C:MDL D:DF E:MEAN F:Geomean G:MEDIAN H:MIN I:MAX
    # J:5TH K:25TH L:75TH M:95TH N:空 O:说明 P~CX:样品数据(88列)
    sample_start = 16  # P列
    last_sample = sample_start + n_s - 1

    blanks_name = text['blank']
    bottle_name = text['bottle']
    al = blank_info['avg_l']
    ml = blank_info['mdl_l']
    hl = blank_info['half_l']

    # Row 1
    ws.cell(row=1, column=1, value=text['compound']); sty(ws.cell(row=1,column=1), S['hdr'])
    ws.cell(row=1, column=2, value='bottle'); sty(ws.cell(row=1,column=2), S['hdr'])
    ws.cell(row=1, column=3, value='bottle'); sty(ws.cell(row=1,column=3), S['hdr'])
    for c in range(4, 14): ws.cell(row=1, column=c, value=None)
    ws.cell(row=1, column=14, value=None)
    ws.cell(row=1, column=15, value=None)

    # 样品列头
    for i, (_, _, hdr) in enumerate(sample_cols):
        ws.cell(row=1, column=sample_start+i, value=hdr.replace('0527 Urine-',''))
        sty(ws.cell(row=1,column=sample_start+i), S['hdr'])

    # Row 2
    ws.cell(row=2, column=1, value=text['group']); sty(ws.cell(row=2,column=1), S['hdr'])
    ws.cell(row=2, column=2, value='BLANK average'); sty(ws.cell(row=2,column=2), S['hdr'])
    ws.cell(row=2, column=3, value='MDL'); sty(ws.cell(row=2,column=3), S['hdr'])

    stat_names = ['DF 检出率','MEAN','Geometric Mean','MEDIAN','MIN','MAX',
                  '5TH','25TH','75TH','95TH']
    for i, nm in enumerate(stat_names):
        ws.cell(row=2, column=4+i, value=nm)
        sty(ws.cell(row=2,column=4+i), S['stat'])

    ws.cell(row=2, column=15, value=text['sample_volume'])
    sty(ws.cell(row=2,column=15), S['yell'])

    for i in range(n_s):
        ws.cell(row=2, column=sample_start+i, value=cfg.get('sample_volume_ml', 2))
        sty(ws.cell(row=2,column=sample_start+i), S['data'])

    # 统计列 D-M (cols 4-13): Row1+Row2 合并
    stat_end = 4 + len(stat_names) - 1  # col 13
    for col in range(4, stat_end + 1):
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        # 合并后，合并区域的左上角单元格保留值，其他被清除
        # 把 row2 的值移到 row1
        cl = get_column_letter(col)
        merged_cell = ws.cell(row=1, column=col)
        merged_cell.value = stat_names[col - 4]
        sty(merged_cell, S['stat'])

    # Row 3 单位
    ws.cell(row=3, column=1, value=cfg['sample_type']); sty(ws.cell(row=3,column=1), S['cmpd'])
    ws.cell(row=3, column=2, value='ng/mL'); sty(ws.cell(row=3,column=2), S['hdr'])
    ws.cell(row=3, column=3, value='ng/mL'); sty(ws.cell(row=3,column=3), S['hdr'])
    ws.cell(row=3, column=4, value='%'); sty(ws.cell(row=3,column=4), S['stat'])
    for c in range(5, 14):
        ws.cell(row=3, column=c, value=unit); sty(ws.cell(row=3,column=c), S['stat'])

    # A38 标注, B38 存储换算因子
    ws.cell(row=38, column=1, value='换算因子')
    sty(ws.cell(row=38,column=1), S['bold'])
    ws.cell(row=38, column=2, value=cf)
    ws.cell(row=38,column=2).number_format = '0.000000'
    sty(ws.cell(row=38,column=2), S['data'])

    # 行号映射
    s2_first = blank_info['first_row']
    target_compounds = cfg.get('target_compounds', TARGET_COMPS)
    comp2row_s2 = {c: s2_first + i for i, c in enumerate(target_compounds)}
    comp2row_s3 = {c: s3_first + i for i, c in enumerate(target_compounds)}

    row = 4
    for comp in target_compounds:
        analysis = cfg['_analysis_results'][comp]
        ws.cell(row=row, column=1, value=comp); sty(ws.cell(row=row,column=1), S['cmpd'])
        s2r = comp2row_s2.get(comp)
        s3r = comp2row_s3.get(comp)

        # B: BLANK average
        cb = ws.cell(row=row, column=2)
        if s2r: cb.value = f"='{blanks_name}'!{al}{s2r}"; cb.number_format = '0.000000'
        sty(cb, S['data'])
        cache_formula_value(cfg, ws, cb, analysis['blank_average'])

        # C: MDL in the report/sample unit (bottle MDL is converted once).
        cc = ws.cell(row=row, column=3)
        if s2r: cc.value = f"='{blanks_name}'!{ml}{s2r}*$B$38"; cc.number_format = '0.000000'
        sty(cc, S['data'])
        cache_formula_value(cfg, ws, cc, analysis['report_mdl'])

        # Final concentration range (P onward). DF now follows the old template
        # directly from this range, so no hidden detection-status columns are needed.
        sl = get_column_letter(sample_start)
        el = get_column_letter(last_sample)
        sr = f'{sl}{row}:{el}{row}'
        # D: DF 检出率，恢复旧模板规则：有数值结果数 ÷ 全部样品列数。
        cd = ws.cell(row=row, column=4)
        cd.value = f'=COUNT({sr})/COLUMNS({sr})'
        cd.number_format = '0.00%'
        sty(cd, S['stat'])
        cache_formula_value(cfg, ws, cd, analysis['df_fraction'])

        # E~I: 旧模板规则，DF>50%才计算，否则显示NC。
        funcs = {5:'AVERAGE', 6:'GEOMEAN', 7:'MEDIAN', 8:'MIN', 9:'MAX'}
        statistic_keys = {5:'mean', 6:'geomean', 7:'median', 8:'min', 9:'max'}
        for col, func in funcs.items():
            c = ws.cell(row=row, column=col)
            c.value = f'=IF(D{row}>50%,{func}({sr}),"NC")'
            c.number_format = '0.000000'
            sty(c, S['stat'])
            cache_formula_value(cfg, ws, c, analysis[statistic_keys[col]] if analysis['statistics_eligible'] else 'NC')

        # J~M: Percentiles
        for col, pct, key in [(10,0.05,'p05'),(11,0.25,'p25'),(12,0.75,'p75'),(13,0.95,'p95')]:
            c = ws.cell(row=row, column=col)
            c.value = f'=IF(D{row}>50%,PERCENTILE({sr},{pct}),"NC")'
            c.number_format = '0.000000'
            sty(c, S['stat'])
            cache_formula_value(cfg, ws, c, analysis[key] if analysis['statistics_eligible'] else 'NC')

        # 样品数据列 P~CX
        for i, (_, _, sample_header) in enumerate(sample_cols):
            col = sample_start + i
            s3_cl = get_column_letter(2 + i)

            if s2r and s3r:
                formula = (
                    f"=IF('{bottle_name}'!{s3_cl}{s3r}=\"\",\"\","
                    f"IF('{bottle_name}'!{s3_cl}{s3r}>='{blanks_name}'!{ml}{s2r},"
                    f"('{bottle_name}'!{s3_cl}{s3r}-'{blanks_name}'!{al}{s2r})*$B$38,"
                    f"'{blanks_name}'!{hl}{s2r}))"
                )
                ws.cell(row=row, column=col).value = formula
                ws.cell(row=row, column=col).number_format = '0.000000'
                cache_formula_value(cfg, ws, ws.cell(row=row, column=col), analysis['sample_values'].get(sample_header))
            sty(ws.cell(row=row, column=col), S['data'])
        row += 1

    ws.row_dimensions[1].height = 19.5
    ws.row_dimensions[2].height = 17.25
    ws.column_dimensions['A'].width = 30.0
    return ws


def build_summary_sheet(wb, final_sheet, blank_info, sample_cols, S, cfg):
    """Create a compact, formula-driven descriptive-statistics sheet."""
    text = workbook_text(cfg)
    title = text['summary']
    ws = wb.create_sheet(title)
    final_name = final_sheet.title
    blanks_name = text['blank']
    sample_start = 16
    last_sample = sample_start + len(sample_cols) - 1
    first_letter = get_column_letter(sample_start)
    last_letter = get_column_letter(last_sample)
    target_compounds = cfg.get('target_compounds', TARGET_COMPS)
    headers = text['summary_headers']
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value=text['summary_title'])
    sty(ws.cell(row=1, column=1), S['hdr'])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(row=2, column=1, value=text['summary_note'])
    sty(ws.cell(row=2, column=1), S['yell'])
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
        sty(ws.cell(row=3, column=col), S['hdr'])

    mdl_letter = blank_info['mdl_l']
    for index, compound in enumerate(target_compounds):
        analysis = cfg['_analysis_results'][compound]
        row = 4 + index
        final_row = 4 + index
        sample_range = f"'{final_name}'!{first_letter}{final_row}:{last_letter}{final_row}"
        ws.cell(row=row, column=1, value=compound)
        ws.cell(row=row, column=2, value=compound_metadata_for(compound, cfg.get('compound_metadata'))['chain_length'])
        # DF is displayed directly as a percentage number (for example 33.3),
        # consistent with the online preview and the "DF (%)" heading.
        df_cell = ws.cell(row=row, column=3, value=f'={significant_digits_formula(f"\'{final_name}\'!D{final_row}*100")}')
        cache_formula_value(cfg, ws, df_cell, _round_significant(analysis['df_fraction'] * 100))
        # Old template rule: descriptive statistics are displayed only at DF>50%.
        median_formula = significant_digits_formula(f'MEDIAN({sample_range})')
        q1_formula = significant_digits_formula(f'PERCENTILE({sample_range},0.25)')
        q3_formula = significant_digits_formula(f'PERCENTILE({sample_range},0.75)')
        median_cell = ws.cell(row=row, column=4, value=(
            f'=IF(\'{final_name}\'!D{final_row}>50%,{median_formula}&" ("&{q1_formula}&"-"&{q3_formula}&")","NC")'
        ))
        median_iqr = (
            f'{_format_significant(analysis["median"])} '
            f'({_format_significant(analysis["p25"])}-{_format_significant(analysis["p75"])})'
            if analysis['statistics_eligible'] else 'NC'
        )
        cache_formula_value(cfg, ws, median_cell, median_iqr)
        blank_row = blank_info['row_map'].get(compound)
        if blank_row is None:
            raise ValueError(f'{compound}: missing MDL row in blank sheet.')
        mdl_ref = f"'{blanks_name}'!{mdl_letter}{blank_row}"
        # Use the same report-unit MDL as Final. conc; blank-zero analytes use
        # the explicit calibration/SN formula and are converted exactly once.
        blank_range = f"'{blanks_name}'!B{blank_row}:{get_column_letter(1 + cfg.get('blank_column_count', 2))}{blank_row}"
        report_mdl_formula = mdl_report_formula(compound, mdl_ref, cfg)
        report_mql_formula = mql_report_formula(compound, mdl_ref, blank_range, cfg)
        mdl_cell = ws.cell(row=row, column=5, value=f'={significant_digits_formula(report_mdl_formula[1:])}')
        mql_cell = ws.cell(row=row, column=6, value=f'={significant_digits_formula(report_mql_formula[1:])}')
        cache_formula_value(cfg, ws, mdl_cell, _round_significant(analysis['report_mdl']))
        cache_formula_value(cfg, ws, mql_cell, _round_significant(analysis['report_mql']))
        for col in range(1, 7):
            sty(ws.cell(row=row, column=col), S['cmpd'] if col in (1, 2) else S['stat'])
        # Each formula has already rounded the result to three significant
        # digits. General avoids a fixed six-decimal display when users paste
        # the compact descriptive statistics into manuscripts.
        for col in (3, 5, 6):
            ws.cell(row=row, column=col).number_format = 'General'

    for col, width in enumerate([30, 12, 12, 28, 14, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:F{3 + len(target_compounds)}'
    return ws


# ============================================================
# 说明书
# ============================================================
def build_info_sheet(wb, S, cfg):
    text = workbook_text(cfg)
    ws = wb.create_sheet(text['notes'])
    english = cfg.get('language') == 'en'
    rows = [
        [text['notes_title']],
        ['Area', 'Formula/rule', 'Source', 'Note'] if english else ['区域', '公式/规则', '来源', '说明'],
        ['Matrix spike measured values', 'Copy the matching raw MS concentration', 'Original MS columns', 'MS columns are detected dynamically.' if english else 'matrix spike_1、2……直接复制原始MS实测浓度，列数随文件自动变化。'],
        ['Target/SS recovery', 'measured MS ÷ corresponding theoretical MS spike × 100%', 'Raw MS + user spike settings', 'SS uses its own compound-by-MS theoretical spike; IS has no recovery.' if english else '目标物使用对应MS理论加标浓度；SS使用用户填写的自身逐MS理论加标浓度；IS不计算回收率。'],
        ['Recovery average / SD / SE', 'ROUND(AVERAGE,0); ROUND(STDEV,0); ROUND(SD/SQRT(COUNT),0)', 'Valid recovery cells', 'All three statistics use only numeric recovery cells.' if english else 'average为算术平均；SD为样本标准差；SE=SD÷√有效回收率数量，均四舍五入取整。'],
        ['Blank average', 'AVERAGE(valid blank values); C/S/N path = 0', 'Original blank columns', 'Empty cells are excluded and are not converted to zero.' if english else 'Blank存在非零值时取有效数值平均；C/S/N路径按0；原始空单元格不作为0。'],
        ['MDL (blank not zero)', 'AVERAGE(blank)+TINV(0.02,n-1)×STDEV(blank)', 'Valid procedural blank replicates', 'Equivalent to one-sided 99% t; no division by SQRT(n).' if english else 'TINV(0.02,n−1)等价于单侧99% t值；使用Blank样本标准差，不除以√n。'],
        ['MDL (blank zero)', '3 × calibration concentration ÷ S/N', 'User-entered calibration point and S/N', 'Blank cells are missing, not zero.' if english else '空单元格为缺失值，不等于0。'],
        ['Half sample MDL', 'vial MDL ÷ 2 × conversion factor', 'Blank MDL + conversion factor', 'Used for valid non-detect substitution and already in the final reporting unit.' if english else '用于有效未检出样品替代，结果已经是最终样本报告单位。'],
        ['Final sample concentration', 'empty→empty; vial≥MDL→(vial-blank mean)×factor; otherwise half sample MDL', 'Vial concentration + Blank MDL', 'The source vial value is compared with vial MDL before conversion.' if english else '原始空值保持空；瓶内值≥瓶内MDL时扣Blank并换算；低于MDL但有值时使用1/2样本MDL。'],
        ['DF', 'COUNT(final sample results) ÷ COLUMNS(all sample columns)', 'Final concentration sample range', 'Numeric half-MDL substitutes count; missing source cells remain in the denominator.' if english else '有数值最终浓度数÷全部样品列数；1/2 MDL替代值计数，原始空单元格仍占分母。'],
        ['MEAN / Geometric Mean', 'IF(DF>50%,AVERAGE/GEOMEAN(final range),"NC")', 'Final concentration sample range', 'GEOMEAN requires all participating numeric results to be positive.' if english else 'DF严格大于50%才计算；MEAN为算术平均，Geometric Mean要求参与数值均大于0。'],
        ['MEDIAN / MIN / MAX', 'IF(DF>50%,MEDIAN/MIN/MAX(final range),"NC")', 'Final concentration sample range', 'DF=50% still returns NC.' if english else 'DF严格大于50%才计算；DF等于50%仍显示NC。'],
        ['5TH / 25TH / 75TH / 95TH', 'IF(DF>50%,PERCENTILE(final range,p),"NC")', 'Final concentration sample range', 'p=0.05, 0.25, 0.75 and 0.95; 25TH=Q1 and 75TH=Q3.' if english else '分别使用0.05、0.25、0.75、0.95；25TH=Q1，75TH=Q3；DF≤50%显示NC。'],
        ['Descriptive summary', 'DF×100; Median(Q1-Q3) only if DF>50%; MDL/MQL to 3 significant figures', 'Shared Python results + retained formulas', 'Online preview and downloaded workbook use the same numeric results.' if english else 'DF转百分数；DF>50%才展示Median(Q1-Q3)；MDL/MQL保留3位有效数字；网页与下载使用同一结果。'],
        ['MQL', '(blank mean+10×STDEV(blank))×factor; C/S/N path: 10×C÷S/N×factor', 'Same blank inputs as MDL', 'Reported in final sample units.' if english else 'Blank非零路径为平均值+10倍样本SD；C/S/N路径为10×C÷S/N；最后乘换算因子。'],
        ['IS measured concentrations', 'Copied per IS and per MS cell; no recovery calculation', 'Original unprocessed MS columns', 'IS correction is controlled only by the IS-corrected setting.' if english else 'IS实测浓度由原始未处理表对应MS列自动提取；是否IS校正仍只由界面“数据是否经过IS校正”控制。'],
        ['Parameters', f'sample={cfg["sample_type"]}; sample volume={cfg["sample_volume_ml"]} mL; final volume={cfg["final_volume_ml"]} mL; conversion factor={cfg["conversion_factor"]}', '', ''],
    ]
    is_corrected = bool(cfg.get('is_corrected', False))
    ms_headers = cfg.get('matrix_spike_headers') or []
    if cfg.get('is_compounds'):
        rows.append([])
        rows.append([text['is_additions'], *ms_headers, 'IS corrected' if english else '是否IS校正'])
        for name in cfg.get('is_compounds', []):
            raw_row = cfg.get('_raw_data', {}).get(name, {})
            ms_columns = cfg.get('_matrix_spike_columns') or []
            rows.append([name, *[safe_float(raw_row.get(column_letter)) for _, column_letter, _ in ms_columns], 'yes' if is_corrected else 'no'])
    for r, rd in enumerate(rows, 1):
        for c, val in enumerate(rd, 1):
            cell = ws.cell(row=r, column=c, value=val)
            sty(cell, S['hdr'] if r <= 2 else S['data'])
    for c, w in zip(range(1, max(5, len(ms_headers) + 3)), [35,65,20,60,20,20,20]):
        ws.column_dimensions[get_column_letter(c)].width = w
    return ws


def export_csv_bytes(raw_data, blank_cols, sample_cols, cfg):
    """Create a formula-free CSV containing summary and final concentrations."""
    preview_cfg = dict(cfg)
    preview_cfg['_raw_data'] = raw_data
    english = cfg.get('language') == 'en'
    if english:
        summary_headers = ['Name', 'Chain length', 'DF (%)', 'Median (Q1-Q3)', 'MDL', 'MQL']
        final_title = 'Final concentration'
        is_record_title = 'IS measured concentrations (from raw MS columns)'
        is_status_title = 'IS correction applied'
        yes_no = ('yes', 'no')
    else:
        summary_headers = ['名称', '链长', 'DF (%)', 'Median (Q1-Q3)', 'MDL', 'MQL']
        final_title = '最终浓度'
        is_record_title = 'IS 实测浓度（来源于原始MS列）'
        is_status_title = '是否经过 IS 校正'
        yes_no = ('是', '否')

    rows = [summary_headers]
    for item in compute_preview_summary(raw_data, blank_cols, sample_cols, preview_cfg):
        rows.append([item.get('名称'), item.get('链长'), item.get('DF (%)'), item.get('Median (Q1-Q3)'), item.get('MDL'), item.get('MQL')])
    sample_headers = [header for _, _, header in sample_cols]
    rows.append([])
    rows.append([final_title] + sample_headers)
    for item in compute_preview_final_table(raw_data, blank_cols, sample_cols, preview_cfg):
        name = item.get('名称')
        rows.append([name] + [item.get(header) for header in sample_headers])
    rows.append([])
    ms_headers = cfg.get('matrix_spike_headers') or []
    rows.append([is_record_title, *ms_headers, is_status_title])
    ms_columns = cfg.get('_matrix_spike_columns') or []
    for name in cfg.get('is_compounds', []):
        values = [safe_float(raw_data.get(name, {}).get(column_letter)) for _, column_letter, _ in ms_columns]
        rows.append([name, *values, yes_no[0] if cfg.get('is_corrected', False) else yes_no[1]])
    text = io.StringIO()
    csv.writer(text, lineterminator='\n').writerows(rows)
    return text.getvalue().encode('utf-8-sig')


# ============================================================
# 主处理函数 (可被 Streamlit 调用)
# ============================================================
def process(config=None, return_bytes=False):
    """
    执行完整的数据处理流程。

    Args:
        config: 配置字典，None 则使用默认 CONFIG
        return_bytes: True 返回 bytes (供 Streamlit), False 保存到文件

    Returns:
        如果 return_bytes=True，返回 (bytes, filename)
        否则返回 output_filepath
    """
    cfg = {**CONFIG, **(config or {})}
    input_src = cfg.get('input_file', '')
    input_bytes = cfg.get('input_bytes', None)
    src = input_bytes if input_bytes else input_src
    if not src:
        raise ValueError('No input file or data provided')
    raw_data, blanks, mss, samps, detected_target, detected_is, detected_ss, detected_all = read_raw(src)
    cfg['matrix_spike_headers'] = [header for _, _, header in mss]
    cfg['_matrix_spike_columns'] = mss
    cfg['_raw_data'] = raw_data
    target, is_c, ss_c, all_c = configured_compound_lists(cfg, detected_all, detected_is, detected_ss)
    cfg['target_compounds'] = target
    cfg['is_compounds'] = is_c
    cfg['ss_compounds'] = ss_c
    cfg['all_compounds'] = all_c
    cfg['blank_column_count'] = len(blanks)
    layout_report = validate_input_layout(blanks, mss, samps, target, is_c, ss_c)
    if not layout_report['ready']:
        raise ValueError('Input layout is not processable: ' + '; '.join(layout_report['errors']))
    validate_blank_zero_configuration(raw_data, blanks, target, cfg)
    cfg['_analysis_results'] = compute_analysis_results(raw_data, blanks, samps, cfg)
    cfg['_formula_cache'] = {}

    # 验证
    missing = [c for c in all_c if c not in raw_data]
    if missing:
        print(f"WARNING: {len(missing)} compounds not found: {missing}")

    S = make_styles()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("[1/6] Matrix spike...")
    build_sheet1(wb, raw_data, mss, S, cfg)

    print("[2/6] Blanks_MDL...")
    _, binfo = build_sheet2(wb, raw_data, blanks, S, cfg)

    print("[3/6] Conc. in bottle...")
    _, s3_first = build_sheet3(wb, raw_data, samps, S, cfg)

    print("[4/6] Final conc...")
    final_sheet = build_sheet4(wb, raw_data, samps, binfo, s3_first, S, cfg)

    print("[5/6] Descriptive summary...")
    build_summary_sheet(wb, final_sheet, binfo, samps, S, cfg)

    print("[6/6] Info sheet...")
    build_info_sheet(wb, S, cfg)

    if return_bytes:
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        xlsx_bytes = inject_formula_cached_values(output.getvalue(), cfg['_formula_cache'])
        wb.close()
        if cfg.get('output_format', 'xlsx').lower() == 'csv':
            return export_csv_bytes(raw_data, blanks, samps, cfg), cfg.get('output_file', 'processed_data.csv').replace('.xlsx', '.csv')
        return xlsx_bytes, cfg.get('output_file', 'processed_data.xlsx')
    else:
        out = cfg['output_file']
        print(f"Saving to: {out}")
        output = io.BytesIO()
        wb.save(output)
        wb.close()
        with open(out, 'wb') as destination:
            destination.write(inject_formula_cached_values(output.getvalue(), cfg['_formula_cache']))
        print("Done!")
        return out


# ============================================================
# 入口
# ============================================================
if __name__ == '__main__':
    process()
